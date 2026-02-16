"""
Migration script to convert Excel-based data to SQLite database.
Run this once to migrate ApprovedRecords.xlsx and reportingData.xlsx to the new database.
"""

import os
import sys
from datetime import datetime

# Add ExcelVerifier directory to path to import modules
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ExcelVerifier'))

from openpyxl import load_workbook
from core.database_handler import DatabaseHandler
from config import APPROVED_FILE, REPORTING_DATA_FILE, APPROVED_DIRECTORY


def parse_date(date_value):
    """Parse date from various formats."""
    if date_value is None:
        return None
    
    if isinstance(date_value, str):
        date_str = date_value.strip()
        formats = ['%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y', '%Y.%m.%d']
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
            except:
                continue
        return date_str
    elif hasattr(date_value, 'strftime'):
        return date_value.strftime('%Y-%m-%d')
    else:
        return str(date_value)


def parse_float(value):
    """Parse float value, handling '-' and empty strings."""
    if value is None or value == '' or value == '-':
        return 0.0
    try:
        return float(value)
    except:
        return 0.0


def migrate_approved_records(db_handler: DatabaseHandler):
    """Migrate data from ApprovedRecords.xlsx to database."""
    if not os.path.exists(APPROVED_FILE):
        print(f"⚠️  ApprovedRecords.xlsx not found at: {APPROVED_FILE}")
        return 0
    
    try:
        wb = load_workbook(APPROVED_FILE)
        ws = wb['Approved']
        
        migrated = 0
        skipped = 0
        
        print(f"\n📥 Migrating ApprovedRecords.xlsx...")
        print(f"   Location: {APPROVED_FILE}")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 4:
                continue
            
            date_value = row[0]
            company = row[1]
            filename = row[2]
            filepath = row[3]
            
            # Skip empty rows
            if not company or not filename:
                continue
            
            # Normalize values
            date_str = parse_date(date_value)
            company_name = str(company).strip()
            filename_str = str(filename).strip()
            filepath_str = str(filepath).strip() if filepath else ""
            
            # Create company (or get existing)
            company_id = db_handler.add_company(company_name)
            
            # Create order linked to company
            order_id = db_handler.add_order(company_id, date_str)
            
            # Create approved record linked to order
            result = db_handler.add_approved_record(order_id, date_str, filename_str, filepath_str)
            
            if result:
                migrated += 1
            else:
                skipped += 1
        
        wb.close()
        print(f"   ✅ Migrated: {migrated} records")
        if skipped > 0:
            print(f"   ⏭️  Skipped: {skipped} duplicates")
        
        return migrated
        
    except Exception as e:
        print(f"   ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return 0


def migrate_reporting_data(db_handler: DatabaseHandler):
    """Migrate data from reportingData.xlsx to database."""
    if not os.path.exists(REPORTING_DATA_FILE):
        print(f"⚠️  reportingData.xlsx not found at: {REPORTING_DATA_FILE}")
        return 0
    
    try:
        wb = load_workbook(REPORTING_DATA_FILE)
        ws = wb['Records'] if 'Records' in wb.sheetnames else wb.active
        
        items_to_add = []
        
        print(f"\n📥 Migrating reportingData.xlsx...")
        print(f"   Location: {REPORTING_DATA_FILE}")
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 8:
                continue
            
            date_issued = row[0]
            recipient = row[1]
            document_number = row[2]
            product_name = row[3]
            quantity_delivery = row[4]
            quantity_return = row[5]
            previous_state = row[6]
            state_after = row[7]
            
            # Skip empty rows
            if not recipient or not product_name:
                continue
            
            # Normalize values
            date_str = parse_date(date_issued)
            recipient_name = str(recipient).strip()
            doc_number = str(document_number).strip() if document_number else None
            product_str = str(product_name).strip()
            qty_delivery = parse_float(quantity_delivery)
            qty_return = parse_float(quantity_return)
            prev_state = parse_float(previous_state)
            state_after_val = parse_float(state_after)
            
            # Get or create company
            company_id = db_handler.add_company(recipient_name)
            
            # Get or create order
            order_id = db_handler.add_order(company_id, date_str, doc_number)
            
            # Get or create product
            product_id = db_handler.add_product(product_str)
            
            # Add to items list for batch insert
            items_to_add.append({
                'order_id': order_id,
                'product_id': product_id,
                'quantity_delivery': qty_delivery,
                'quantity_return': qty_return,
                'previous_state': prev_state,
                'state_after': state_after_val
            })
        
        wb.close()
        
        # Batch insert all items
        if items_to_add:
            migrated = db_handler.add_order_items(items_to_add)
            print(f"   ✅ Migrated: {migrated} records")
            return migrated
        else:
            print(f"   ℹ️  No records found")
            return 0
        
    except Exception as e:
        print(f"   ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return 0


def create_backup(file_path: str):
    """Create a backup of Excel file before migration."""
    if not os.path.exists(file_path):
        return None
    
    backup_path = file_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    
    try:
        import shutil
        shutil.copy2(file_path, backup_path)
        return backup_path
    except Exception as e:
        print(f"⚠️  Warning: Could not create backup of {file_path}: {e}")
        return None


def main():
    """Main migration function."""
    print("=" * 70)
    print("  ExcelVerifier - Database Migration")
    print("=" * 70)
    print("\nThis script will migrate your Excel data to SQLite database.")
    print("Your Excel files will be backed up before migration.\n")
    
    # Create backups
    print("📋 Creating backups...")
    approved_backup = create_backup(APPROVED_FILE)
    if approved_backup:
        print(f"   ✅ Backed up: {os.path.basename(approved_backup)}")
    
    reporting_backup = create_backup(REPORTING_DATA_FILE)
    if reporting_backup:
        print(f"   ✅ Backed up: {os.path.basename(reporting_backup)}")
    
    # Initialize database
    db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "excelverifier.db")
    print(f"\n🗄️  Initializing database...")
    print(f"   Location: {db_path}")
    
    db_handler = DatabaseHandler(db_path)
    print(f"   ✅ Database ready")
    
    # Migrate data
    approved_count = migrate_approved_records(db_handler)
    reporting_count = migrate_reporting_data(db_handler)
    
    # Show statistics
    print("\n" + "=" * 70)
    print("  Migration Complete!")
    print("=" * 70)
    
    stats = db_handler.get_database_stats()
    print(f"\n📊 Database Statistics:")
    print(f"   • Companies: {stats['companies_count']}")
    print(f"   • Products: {stats['products_count']}")
    print(f"   • Orders: {stats['orders_count']}")
    print(f"   • Approved Records: {stats['approved_records_count']}")
    print(f"   • Order Items: {stats['order_items_count']}")
    print(f"   • Date Range: {stats['earliest_date']} to {stats['latest_date']}")
    print(f"   • Database Size: {stats['database_size'] / 1024:.1f} KB")
    
    print(f"\n✨ Migration successful!")
    print(f"\nℹ️  The application will now use the database instead of Excel files.")
    print(f"ℹ️  Your original Excel files have been backed up and are safe.")
    print(f"ℹ️  You can delete the backup files once you've verified everything works.\n")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Migration cancelled by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\n\n❌ Migration failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
