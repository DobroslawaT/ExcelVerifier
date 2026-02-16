from openpyxl import load_workbook

latest_report = "Raport_ButloDni_20260205_1443.xlsx"
wb = load_workbook(latest_report)
print(f"Report: {latest_report}")
print(f"Sheets: {wb.sheetnames}")

# Count rows in each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"  {sheet_name}: {ws.max_row} rows, {ws.max_column} columns")
