from __future__ import annotations

import argparse
import os
from datetime import datetime
from calendar import monthrange
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_INPUT_DIR = r"C:\Users\dobro\Documents\Automatyzacja Tomsystems\ExcelVerifier\Reports data"


def extract_nip(text: str | None) -> str | None:
    if not text:
        return None
    text = str(text).strip()
    import re

    formatted_match = re.search(r"(\d{3})-(\d{2})-(\d{2})-(\d{3})", text)
    if formatted_match:
        return f"{formatted_match.group(1)}{formatted_match.group(2)}{formatted_match.group(3)}{formatted_match.group(4)}"

    formatted_match2 = re.search(r"(\d{3})-(\d{3})-(\d{2})-(\d{2})", text)
    if formatted_match2:
        return f"{formatted_match2.group(1)}{formatted_match2.group(2)}{formatted_match2.group(3)}{formatted_match2.group(4)}"

    digits_match = re.search(r"(?:^|\s|[^\d])(\d{10})(?:\s|$|[^\d])", text)
    if digits_match:
        return digits_match.group(1)

    return None


def list_excel_files(root_dir: str) -> List[str]:
    paths: List[str] = []
    extensions = {".xlsx", ".xlsm", ".xls", ".xlsb"}
    for dirpath, _, filenames in os.walk(root_dir):
        for name in filenames:
            if os.path.splitext(name)[1].lower() not in extensions:
                continue
            if name.startswith("~$"):
                continue
            if name.lower().startswith("raport_butlodni_"):
                continue
            paths.append(os.path.join(dirpath, name))
    return paths


def parse_header_date(value) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%d.%m.%Y")
        except Exception:
            return None
    return None


def build_report(input_dir: str, output_dir: str) -> str:
    if not os.path.isdir(input_dir):
        parent_dir = os.path.dirname(os.path.normpath(input_dir))
        if parent_dir and parent_dir != input_dir:
            input_dir = parent_dir

    excel_files = list_excel_files(input_dir)
    if not excel_files:
        raise RuntimeError(
            "No Excel files found in: "
            f"{input_dir} (looking for .xlsx, .xlsm, .xls, .xlsb in subfolders)"
        )

    all_data: List[pd.DataFrame] = []

    for file_path in excel_files:
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            odbiorca = str(ws["B1"].value).strip() if ws["B1"].value else None
            data_wyst = parse_header_date(ws["D1"].value)
            nr_dok = str(ws["F1"].value).strip() if ws["F1"].value else None

            rows = []
            for r in range(4, ws.max_row + 1):
                nazwa = ws.cell(row=r, column=2).value
                stan_po = ws.cell(row=r, column=7).value
                if not nazwa and stan_po is None:
                    continue

                row_dict = {
                    "Odbiorca": odbiorca,
                    "NIP": extract_nip(odbiorca),
                    "Data wystawienia": data_wyst,
                    "Nr dokumentu": nr_dok,
                    "Nazwa": str(nazwa).strip() if nazwa else "Nieokreślony",
                    "Ilość zamówiona": ws.cell(row=r, column=3).value,
                    "Ilość zwrócona": ws.cell(row=r, column=5).value,
                    "stan poprzedni": ws.cell(row=r, column=6).value,
                    "stan po wymianie": stan_po,
                }
                rows.append(row_dict)

            if rows:
                all_data.append(pd.DataFrame(rows))

        except Exception as exc:
            print(f"Błąd przy pliku {file_path}: {exc}")
            continue

    if not all_data:
        raise RuntimeError(f"Brak danych. Przeszukano {len(excel_files)} plików.")

    df = pd.concat(all_data, ignore_index=True)

    df = df.drop_duplicates(
        subset=[
            "Odbiorca",
            "NIP",
            "Nazwa",
            "Data wystawienia",
            "Nr dokumentu",
            "Ilość zamówiona",
            "Ilość zwrócona",
            "stan poprzedni",
            "stan po wymianie",
        ],
        keep="first",
    )

    df["Data wystawienia"] = pd.to_datetime(df["Data wystawienia"], errors="coerce")
    numeric_cols = ["stan po wymianie", "stan poprzedni", "Ilość zamówiona", "Ilość zwrócona"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col].astype(str).str.replace(",", "."), errors="coerce").fillna(0)

    report_end = pd.Timestamp(datetime.now().date())

    min_date = df["Data wystawienia"].min()
    if pd.notna(min_date):
        start_month = pd.Timestamp(year=min_date.year, month=min_date.month, day=1)
        month_starts = pd.date_range(start=start_month, end=report_end, freq="MS")
        synthetic_rows = []

        for (odbiorca, nazwa), group in df.groupby(["Odbiorca", "Nazwa"]):
            group_sorted = group.sort_values("Data wystawienia")
            existing_months = set(group_sorted["Data wystawienia"].dt.strftime("%Y-%m"))

            for ms in month_starts:
                month_key = ms.strftime("%Y-%m")
                if month_key in existing_months:
                    continue
                prior = group_sorted[group_sorted["Data wystawienia"] < ms]
                if prior.empty:
                    continue
                last_row = prior.iloc[-1]

                synthetic = last_row.copy()
                synthetic["Data wystawienia"] = ms
                synthetic["Ilość zamówiona"] = 0
                synthetic["Ilość zwrócona"] = 0
                synthetic["stan poprzedni"] = last_row["stan po wymianie"]
                synthetic["stan po wymianie"] = last_row["stan po wymianie"]
                synthetic_rows.append(synthetic)

        if synthetic_rows:
            df = pd.concat([df, pd.DataFrame(synthetic_rows)], ignore_index=True)

    df = df.drop_duplicates(
        subset=[
            "Odbiorca",
            "NIP",
            "Nazwa",
            "Data wystawienia",
            "Nr dokumentu",
            "Ilość zamówiona",
            "Ilość zwrócona",
            "stan poprzedni",
            "stan po wymianie",
        ],
        keep="first",
    )

    if df.empty:
        raise RuntimeError("Brak danych po przefiltrowaniu dat.")

    df.sort_values(
        by=["Odbiorca", "Nazwa", "Data wystawienia"],
        ascending=[True, True, True],
        inplace=True,
    )
    df["Data następna"] = df.groupby(["Odbiorca", "Nazwa"])["Data wystawienia"].shift(-1)
    today = pd.Timestamp(datetime.now().date())
    df["Data następna"] = df["Data następna"].fillna(today)
    df["Liczba dni"] = (df["Data następna"] - df["Data wystawienia"]).dt.days
    df["Liczba dni"] = df["Liczba dni"].apply(lambda x: x if x >= 0 else 0)
    df["Miesiąc"] = df["Data wystawienia"].dt.strftime("%Y-%m")
    df["Is_first_of_month"] = df.groupby(["Odbiorca", "Nazwa", "Miesiąc"]).cumcount() == 0
    df["Butlo-dni"] = df.apply(
        lambda row: row["stan poprzedni"] * row["Liczba dni"]
        if row["Is_first_of_month"]
        else row["stan po wymianie"] * row["Liczba dni"],
        axis=1,
    )

    raw_cols = [
        "Odbiorca",
        "NIP",
        "Nazwa",
        "Nr dokumentu",
        "Data wystawienia",
        "Ilość zamówiona",
        "Ilość zwrócona",
        "stan poprzedni",
        "stan po wymianie",
        "Miesiąc",
    ]
    df_raw = df[raw_cols].copy()

    df_raw = df_raw.drop_duplicates(
        subset=[
            "Odbiorca",
            "NIP",
            "Nazwa",
            "Data wystawienia",
            "Nr dokumentu",
            "Ilość zamówiona",
            "Ilość zwrócona",
            "stan poprzedni",
            "stan po wymianie",
        ],
        keep="first",
    )

    calc_rows = []
    for (odbiorca, nazwa, miesiac), group in df.groupby(["Odbiorca", "Nazwa", "Miesiąc"]):
        year, month = map(int, miesiac.split("-"))
        first_day = pd.Timestamp(year=year, month=month, day=1)
        last_day = pd.Timestamp(year=year, month=month, day=monthrange(year, month)[1])

        group = group.sort_values("Data wystawienia").reset_index(drop=True)
        current_start = first_day

        for idx, row in group.iterrows():
            report_date = row["Data wystawienia"]
            if idx == 0:
                days = (report_date - current_start).days + 1
                stan = row["stan poprzedni"]
            else:
                days = (report_date - current_start).days
                stan = group.loc[idx - 1, "stan po wymianie"]

            calc_rows.append(
                {
                    "Odbiorca": odbiorca,
                    "Nazwa": nazwa,
                    "Nr dokumentu": row["Nr dokumentu"] if idx == 0 else group.loc[idx - 1, "Nr dokumentu"],
                    "Data początkowa": current_start,
                    "Data końcowa": report_date,
                    "liczba dni": days,
                    "Stan": stan,
                    "Butlo-dni": stan * days if days > 0 else 0,
                    "Miesiąc": miesiac,
                }
            )
            current_start = report_date

        last_row = group.iloc[-1]
        days_end = (last_day - current_start).days
        stan_end = last_row["stan po wymianie"]
        calc_rows.append(
            {
                "Odbiorca": odbiorca,
                "Nazwa": nazwa,
                "Nr dokumentu": last_row["Nr dokumentu"],
                "Data początkowa": current_start,
                "Data końcowa": last_day,
                "liczba dni": days_end,
                "Stan": stan_end,
                "Butlo-dni": stan_end * days_end if days_end > 0 else 0,
                "Miesiąc": miesiac,
            }
        )

    df_calc = pd.DataFrame(calc_rows)

    output_filename = f"Raport_ButloDni_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, output_filename)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_raw.to_excel(writer, sheet_name="Dane Szczegółowe", index=False)
        df_calc.to_excel(writer, sheet_name="Obliczenia", index=False)

        df_calc_pos = df_calc.copy()
        df_calc_pos["Butlo-dni"] = df_calc_pos["Butlo-dni"].apply(lambda x: max(0, x))
        sum_butlo = df_calc_pos.groupby(["Odbiorca", "Nazwa", "Miesiąc"], as_index=False)[["Butlo-dni"]].sum()
        sum_rot = df_raw.groupby(["Odbiorca", "Nazwa", "Miesiąc"], as_index=False)[["Ilość zwrócona"]].sum()
        summary = sum_butlo.merge(sum_rot, on=["Odbiorca", "Nazwa", "Miesiąc"], how="outer")
        nip_map = df[["Odbiorca", "NIP"]].drop_duplicates()
        summary = summary.merge(nip_map, on="Odbiorca", how="left")
        summary = summary[["Odbiorca", "NIP", "Nazwa", "Miesiąc", "Butlo-dni", "Ilość zwrócona"]]
        summary = summary.rename(columns={"Ilość zwrócona": "rotacja"})
        summary.to_excel(writer, sheet_name="Podsumowanie", index=False)

        for sn in ["Dane Szczegółowe", "Obliczenia"]:
            ws = writer.sheets[sn]
            for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if sn == "Dane Szczegółowe":
                    r[3].number_format = "dd.mm.yyyy"
                else:
                    r[3].number_format = "dd.mm.yyyy"
                    r[4].number_format = "dd.mm.yyyy"

        for sn in writer.sheets:
            ws = writer.sheets[sn]
            for col in ws.columns:
                width = 10
                for cell in col:
                    try:
                        width = max(width, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[get_column_letter(col[0].column)].width = width + 2

    create_pivot(output_path)

    return output_path


def create_pivot(output_path: str) -> None:
    excel = None
    wb_com = None
    try:
        import pythoncom
        import win32com.client as win32
        import time
        import gc

        pythoncom.CoInitialize()
        time.sleep(1.0)

        abs_path = os.path.abspath(output_path)

        xlDatabase = 1
        xlRowField = 1
        xlColumnField = 2
        xlPageField = 3
        xlSum = -4157
        xlUp = -4162
        xlToLeft = -4159

        try:
            excel = win32.DispatchEx("Excel.Application")
        except Exception:
            excel = win32.Dispatch("Excel.Application")

        excel.Visible = False
        excel.DisplayAlerts = False

        wb_com = excel.Workbooks.Open(abs_path)
        ws_src = wb_com.Sheets("Podsumowanie")

        last_row = ws_src.Cells(ws_src.Rows.Count, 1).End(xlUp).Row
        last_col = ws_src.Cells(1, ws_src.Columns.Count).End(xlToLeft).Column
        src_range = ws_src.Range(ws_src.Cells(1, 1), ws_src.Cells(last_row, last_col))

        ws_pivot = wb_com.Sheets.Add()
        ws_pivot.Name = "Pivot"

        pc = wb_com.PivotCaches().Create(SourceType=xlDatabase, SourceData=src_range)
        pt = pc.CreatePivotTable(TableDestination=ws_pivot.Range("A3"), TableName="RaportPivot")

        pt.PivotFields("Odbiorca").Orientation = xlPageField
        pt.PivotFields("NIP").Orientation = xlPageField
        pt.PivotFields("Miesiąc").Orientation = xlPageField

        pt.PivotFields("Nazwa").Orientation = xlRowField

        df1 = pt.AddDataField(pt.PivotFields("Butlo-dni"), "Suma Butlo-dni", xlSum)
        df1.NumberFormat = "0.00"
        df2 = pt.AddDataField(pt.PivotFields("rotacja"), "Suma rotacja", xlSum)
        df2.NumberFormat = "0"

        pt.DataPivotField.Orientation = xlColumnField
        pt.RowAxisLayout(1)

        ws_pivot.Activate()
        wb_com.Save()
        print("Pivot table created successfully.")

    except Exception as exc:
        print(f"Pivot creation failed: {exc}")
    finally:
        if wb_com:
            try:
                wb_com.Close(SaveChanges=True)
            except Exception:
                pass
            wb_com = None

        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
            del excel

        try:
            import pythoncom
            pythoncom.CoUninitialize()
        except Exception:
            pass


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a report from all Excel files in Reports data.")
    parser.add_argument("--input-dir", default=DEFAULT_INPUT_DIR, help="Root directory to scan for Excel files")
    parser.add_argument("--output-dir", default=DEFAULT_INPUT_DIR, help="Where to save the generated report")
    args = parser.parse_args()

    output_path = build_report(args.input_dir, args.output_dir)
    print(f"Report created: {output_path}")


if __name__ == "__main__":
    main()
