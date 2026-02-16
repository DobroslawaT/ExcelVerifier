from openpyxl import load_workbook

reports = [
    "Raport_ButloDni_20260205_1443.xlsx",
    "Raport_ButloDni_20260205_1439.xlsx", 
    "Raport_ButloDni_20260205_1438.xlsx"
]

for report in reports:
    try:
        wb = load_workbook(report)
        sheets = wb.sheetnames
        print(f"{report}")
        print(f"  Sheets: {sheets}")
        if 'Pivot' in sheets:
            pivot_ws = wb['Pivot']
            data_rows = 0
            for row in pivot_ws.iter_rows(values_only=True):
                if any(row):
                    data_rows += 1
            print(f"  Pivot has {data_rows} rows with data")
    except Exception as e:
        print(f"  Error: {e}")
    print()
