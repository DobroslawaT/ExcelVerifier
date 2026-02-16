#!/usr/bin/env python
"""Test date extraction from Excel cell D1"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from datetime import datetime

# Find an unapproved report to test with
reports_dir = "Reports"
test_file = None

# Look for Excel files in Reports directory
for company_dir in os.listdir(reports_dir):
    company_path = os.path.join(reports_dir, company_dir)
    if os.path.isdir(company_path):
        for file in os.listdir(company_path):
            if file.endswith('.xlsx'):
                test_file = os.path.join(company_path, file)
                break
    if test_file:
        break

if test_file:
    print(f"Testing with file: {test_file}")
    try:
        wb = load_workbook(test_file)
        ws = wb.active
        date_cell = ws['D1'].value
        
        print(f"\nCell D1 value: {repr(date_cell)}")
        print(f"Cell D1 type: {type(date_cell)}")
        
        # Try to extract month like main_window does
        filter_month = None
        if date_cell:
            if isinstance(date_cell, datetime):
                filter_month = date_cell.strftime('%Y-%m')
                print(f"✓ Parsed as datetime: {filter_month}")
            elif isinstance(date_cell, str):
                try:
                    parsed_date = datetime.strptime(date_cell.strip(), '%d.%m.%Y')
                    filter_month = parsed_date.strftime('%Y-%m')
                    print(f"✓ Parsed as string: {filter_month}")
                except Exception as e:
                    print(f"✗ Failed to parse string: {e}")
        
        if filter_month:
            print(f"\n✓ Filter month would be: {filter_month}")
        else:
            print(f"\n✗ No filter month extracted - would show current month or latest month")
        
        wb.close()
    except Exception as e:
        print(f"Error: {e}")
else:
    print("No Excel files found in Reports directory")
