from openpyxl import load_workbook

wb = load_workbook("Raport_ButloDni_20260205_1443.xlsx")
ws_pivot = wb["Pivot"]

print("Pivot Sheet Content (first 15 rows):")
for i, row in enumerate(ws_pivot.iter_rows(values_only=True), 1):
    if i > 15:
        break
    print(f"  Row {i}: {row[:5] if row else row}")  # Show only first 5 cols

# Check if it's a real pivot table
print(f"\nPivot tables in sheet: {len(ws_pivot._tables)}")
for table in ws_pivot._tables.values():
    print(f"  Table: {table.name}")

# Check for pivot table definition
print(f"\nActive cell: {ws_pivot.active_cell}")
print(f"Max row: {ws_pivot.max_row}, Max col: {ws_pivot.max_column}")
