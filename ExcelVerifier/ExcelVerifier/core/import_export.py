"""
Import/Export functionality for ExcelVerifier.
Allows backing up and restoring data, migrating between computers, and importing from Excel files.
"""

import os
import zipfile
import shutil
import tempfile
import json
from datetime import datetime
from typing import List, Tuple
from openpyxl import load_workbook

from core.database_handler import DatabaseHandler
from core.company_db import load_company_db
from config import DATABASE_FILE, APPROVED_DIRECTORY, REPORTS_ROOT, COMPANY_DB_FILE


class ImportExportHandler:
    """Handles import/export operations for the application."""
    
    def __init__(self):
        self.db = DatabaseHandler(DATABASE_FILE)
    
    def export_all_data(self, output_path: str) -> Tuple[bool, str]:
        """
        Export all application data to a zip file.
        
        Includes:
        - Database (excelverifier.db)
        - All Excel files from Reports folders
        - Company database (company_db.json)
        - Settings (settings.json)
        
        Args:
            output_path: Path where to save the zip file
            
        Returns:
            Tuple of (success, message)
        """
        try:
            # Create temporary directory for staging
            with tempfile.TemporaryDirectory() as temp_dir:
                # Create directory structure
                data_dir = os.path.join(temp_dir, "ExcelVerifier_Data")
                os.makedirs(data_dir, exist_ok=True)
                
                files_added = 0
                
                # 1. Copy database
                if os.path.exists(DATABASE_FILE):
                    shutil.copy2(DATABASE_FILE, os.path.join(data_dir, "excelverifier.db"))
                    files_added += 1
                
                # 2. Export companies from DB to JSON (optional backup)
                companies = load_company_db(COMPANY_DB_FILE)
                if companies:
                    company_json_path = os.path.join(data_dir, "company_db.json")
                    with open(company_json_path, "w", encoding="utf-8") as file_handle:
                        json.dump(companies, file_handle, ensure_ascii=False, indent=2)
                    files_added += 1
                
                # 3. Copy settings
                settings_file = "settings.json"
                if os.path.exists(settings_file):
                    shutil.copy2(settings_file, os.path.join(data_dir, "settings.json"))
                    files_added += 1
                
                # 4. Copy all Excel files and images from Reports directory
                reports_backup = os.path.join(data_dir, "Reports")
                copied_paths = set()
                if os.path.exists(APPROVED_DIRECTORY):
                    # Get all approved records to know which files to backup
                    records = self.db.get_all_approved_records()
                    
                    for record in records:
                        filepath = record['filepath']
                        if os.path.exists(filepath):
                            # Preserve directory structure
                            rel_path = os.path.relpath(filepath, os.path.dirname(APPROVED_DIRECTORY))
                            dest_path = os.path.join(reports_backup, rel_path)
                            if dest_path not in copied_paths:
                                os.makedirs(os.path.dirname(dest_path), exist_ok=True)
                                shutil.copy2(filepath, dest_path)
                                copied_paths.add(dest_path)
                                files_added += 1
                            
                            # Also copy linked image if exists
                            excel_dir = os.path.dirname(filepath)
                            excel_base = os.path.splitext(os.path.basename(filepath))[0]
                            for ext in ['.jpg', '.jpeg', '.png', '.bmp']:
                                img_path = os.path.join(excel_dir, excel_base + ext)
                                if os.path.exists(img_path):
                                    img_dest = os.path.join(reports_backup, os.path.relpath(img_path, os.path.dirname(APPROVED_DIRECTORY)))
                                    if img_dest not in copied_paths:
                                        os.makedirs(os.path.dirname(img_dest), exist_ok=True)
                                        shutil.copy2(img_path, img_dest)
                                        copied_paths.add(img_dest)
                                        files_added += 1
                                    break
                    
                    # Also scan approved directory in case database is empty
                    for root, dirs, files in os.walk(APPROVED_DIRECTORY):
                        for file in files:
                            if not file.lower().endswith(('.xlsx', '.xls')):
                                continue
                            src = os.path.join(root, file)
                            rel_path = os.path.relpath(src, os.path.dirname(APPROVED_DIRECTORY))
                            dest_path = os.path.join(reports_backup, rel_path)
                            if dest_path not in copied_paths:
                                os.makedirs(os.path.dirname(dest_path), exist_ok=True)
                                shutil.copy2(src, dest_path)
                                copied_paths.add(dest_path)
                                files_added += 1
                            
                            base_name = os.path.splitext(file)[0]
                            for ext in ['.jpg', '.jpeg', '.png', '.bmp']:
                                img_path = os.path.join(root, base_name + ext)
                                if os.path.exists(img_path):
                                    img_dest = os.path.join(reports_backup, os.path.relpath(img_path, os.path.dirname(APPROVED_DIRECTORY)))
                                    if img_dest not in copied_paths:
                                        os.makedirs(os.path.dirname(img_dest), exist_ok=True)
                                        shutil.copy2(img_path, img_dest)
                                        copied_paths.add(img_dest)
                                        files_added += 1
                                    break
                
                # 5. Copy unapproved reports
                if os.path.exists(REPORTS_ROOT):
                    for root, dirs, files in os.walk(REPORTS_ROOT):
                        for file in files:
                            if file.endswith('.xlsx'):
                                src = os.path.join(root, file)
                                rel_path = os.path.relpath(src, os.path.dirname(REPORTS_ROOT))
                                dest = os.path.join(data_dir, rel_path)
                                os.makedirs(os.path.dirname(dest), exist_ok=True)
                                shutil.copy2(src, dest)
                                files_added += 1
                                
                                # Copy linked image
                                base_name = os.path.splitext(file)[0]
                                for ext in ['.jpg', '.jpeg', '.png', '.bmp']:
                                    img_src = os.path.join(root, base_name + ext)
                                    if os.path.exists(img_src):
                                        img_dest = os.path.join(data_dir, os.path.relpath(img_src, os.path.dirname(REPORTS_ROOT)))
                                        os.makedirs(os.path.dirname(img_dest), exist_ok=True)
                                        shutil.copy2(img_src, img_dest)
                                        files_added += 1
                                        break
                
                # Create zip file
                shutil.make_archive(output_path.replace('.zip', ''), 'zip', temp_dir)
                
                # Get file size
                zip_size = os.path.getsize(output_path) / (1024 * 1024)  # MB
                
                return True, f"Eksport zakończony!\n\nPliki: {files_added}\nRozmiar: {zip_size:.1f} MB\nLokalizacja: {output_path}"
                
        except Exception as e:
            return False, f"Błąd eksportu: {str(e)}"
    
    def import_all_data(self, zip_path: str, merge: bool = False) -> Tuple[bool, str]:
        """
        Import data from a zip file.
        
        Args:
            zip_path: Path to the zip file to import
            merge: If True, merge with existing data. If False, replace.
            
        Returns:
            Tuple of (success, message)
        """
        try:
            if not os.path.exists(zip_path):
                return False, f"Plik nie istnieje: {zip_path}"
            
            # Extract to temporary directory
            with tempfile.TemporaryDirectory() as temp_dir:
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(temp_dir)
                
                # Find the data directory and relevant files
                data_dir = None
                db_path = None
                company_db_path = None
                reports_dirs = []
                
                for root, dirs, files in os.walk(temp_dir):
                    if os.path.basename(root) == "ExcelVerifier_Data":
                        data_dir = root
                    if not db_path and "excelverifier.db" in files:
                        db_path = os.path.join(root, "excelverifier.db")
                    if not company_db_path and "company_db.json" in files:
                        company_db_path = os.path.join(root, "company_db.json")
                    if "Reports" in dirs:
                        reports_dirs.append(os.path.join(root, "Reports"))
                
                if not data_dir:
                    if db_path:
                        data_dir = os.path.dirname(db_path)
                    elif reports_dirs:
                        data_dir = os.path.dirname(reports_dirs[0])
                
                if not data_dir:
                    return False, "Nieprawidłowy format archiwum"
                
                imported_items = []
                
                # 1. Import database
                if not db_path:
                    db_path = os.path.join(data_dir, "excelverifier.db")
                if db_path and os.path.exists(db_path):
                    if merge:
                        # Merge databases
                        imported_items.append(self._merge_database(db_path))
                    else:
                        # Replace database
                        if os.path.exists(DATABASE_FILE):
                            os.remove(DATABASE_FILE)
                        shutil.copy2(db_path, DATABASE_FILE)
                        imported_items.append("Baza danych zastąpiona")
                
                # 2. Import company database (into DB)
                if not company_db_path:
                    company_db_path = os.path.join(data_dir, "company_db.json")
                if company_db_path and os.path.exists(company_db_path):
                    try:
                        from core.company_db import save_company_db
                        with open(company_db_path, "r", encoding="utf-8") as file_handle:
                            data = json.load(file_handle)
                        if not merge:
                            save_company_db(COMPANY_DB_FILE, data)
                        else:
                            existing = load_company_db(COMPANY_DB_FILE)
                            merged = existing + [item for item in data if item not in existing]
                            save_company_db(COMPANY_DB_FILE, merged)
                        imported_items.append("Baza firm")
                    except Exception:
                        pass
                
                # 3. Import Reports folders
                reports_dirs_to_use = []
                primary_reports_dir = os.path.join(data_dir, "Reports")
                legacy_reports_root = None
                if os.path.exists(os.path.join(data_dir, "Niezatwierdzone")) or os.path.exists(os.path.join(data_dir, "Zatwierdzone")):
                    legacy_reports_root = data_dir
                if os.path.exists(primary_reports_dir):
                    reports_dirs_to_use.append(primary_reports_dir)
                elif legacy_reports_root:
                    reports_dirs_to_use.append(legacy_reports_root)
                if not reports_dirs_to_use and reports_dirs:
                    reports_dirs_to_use = list(dict.fromkeys(reports_dirs))
                
                if reports_dirs_to_use:
                    files_copied = 0
                    files_skipped = 0
                    approved_files = []  # Track approved files for database creation
                    copied_destinations = set()
                    
                    for reports_dir in reports_dirs_to_use:
                        for root, dirs, files in os.walk(reports_dir):
                            for file in files:
                                src = os.path.join(root, file)
                                rel_path = os.path.relpath(src, reports_dir)
                                # rel_path is like "Zatwierdzone/Company/file.xlsx" or "Niezatwierdzone/Company/file.xlsx"
                                # We want to put it in Reports/ directory maintaining this structure
                                rel_parts = rel_path.split(os.sep)
                                if rel_parts and rel_parts[0].lower() == "reports":
                                    rel_parts = rel_parts[1:]
                                if not rel_parts or rel_parts[0] not in ["Zatwierdzone", "Niezatwierdzone"]:
                                    rel_parts = ["Niezatwierdzone"] + rel_parts
                                normalized_rel_path = os.path.join(*rel_parts) if rel_parts else rel_path
                                dest = os.path.join(os.path.dirname(APPROVED_DIRECTORY), normalized_rel_path)
                                os.makedirs(os.path.dirname(dest), exist_ok=True)
                                
                                if dest in copied_destinations:
                                    continue
                                
                                # Don't overwrite existing files if merging
                                if merge and os.path.exists(dest):
                                    files_skipped += 1
                                    if "Zatwierdzone" in normalized_rel_path:
                                        approved_files.append((file, dest))
                                    continue
                                
                                shutil.copy2(src, dest)
                                copied_destinations.add(dest)
                                files_copied += 1
                                
                                # Track approved files for database creation
                                if "Zatwierdzone" in normalized_rel_path:
                                    approved_files.append((file, dest))
                    
                    # Create database records for imported approved files (if database was empty)
                    if approved_files:
                        records_created = self._create_records_for_approved_files(approved_files)
                        if records_created > 0:
                            imported_items.append(f"{records_created} zapisów dla zatwierdzonych plików")
                    
                    imported_items.append(f"{files_copied} plików raportów")
                    if merge and files_skipped > 0:
                        imported_items.append(f"{files_skipped} plików pominięto (już istniały)")

                if merge:
                    disk_approved_files = self._get_approved_files_on_disk()
                    if disk_approved_files:
                        records_created = self._create_records_for_approved_files(disk_approved_files)
                        if records_created > 0:
                            imported_items.append(f"{records_created} zapisów odtworzono z dysku")
                
                return True, f"Import zakończony!\n\nZaimportowano:\n" + "\n".join(f"• {item}" for item in imported_items)
                
        except Exception as e:
            return False, f"Błąd importu: {str(e)}"
    
    def _create_records_for_approved_files(self, approved_files: List[Tuple[str, str]]) -> int:
        """Create database records for approved files if they don't already exist."""
        from datetime import datetime
        
        created = 0
        for filename, filepath in approved_files:
            # Only process Excel files, skip images and other file types
            if not filename.lower().endswith(('.xlsx', '.xls')):
                continue
            
            # Skip if record already exists
            existing = self.db.get_approved_record(filename)
            if existing:
                continue
            
            try:
                # Extract date and company from filename or path
                # Filename format: YYYY-MM-DD_CompanyName.xlsx
                parts = filename.replace('.xlsx', '').replace('.xls', '').split('_', 1)
                date_str = parts[0] if len(parts) > 0 and len(parts[0]) == 10 else datetime.now().strftime('%Y-%m-%d')
                company_name = parts[1] if len(parts) > 1 else os.path.basename(os.path.dirname(filepath))
                
                # Get or create company
                company_id = self.db.add_company(name=company_name, nip=None)
                if not company_id:
                    continue
                
                # Create order
                order_id = self.db.add_order(
                    company_id=company_id,
                    date_issued=date_str,
                    document_number=None
                )
                if not order_id:
                    continue
                
                # Add approved record
                success = self.db.add_approved_record(
                    order_id=order_id,
                    date=date_str,
                    filename=filename,
                    filepath=filepath
                )
                if success:
                    created += 1
            except Exception as e:
                print(f"Error creating record for {filename}: {e}")
                continue
        
        return created

    def _get_approved_files_on_disk(self) -> List[Tuple[str, str]]:
        """Collect approved Excel files currently on disk."""
        approved_files = []
        if not os.path.exists(APPROVED_DIRECTORY):
            return approved_files

        for root, dirs, files in os.walk(APPROVED_DIRECTORY):
            for file in files:
                if not file.lower().endswith((".xlsx", ".xls")):
                    continue
                approved_files.append((file, os.path.join(root, file)))

        return approved_files
    
    def _merge_database(self, source_db_path: str) -> str:
        """Merge records from source database into current database."""
        try:
            # Open source database
            source_db = DatabaseHandler(source_db_path)
            
            # Get all records from source
            approved_records = source_db.get_all_approved_records()
            
            # Import into current database
            added = 0
            skipped = 0
            
            for record in approved_records:
                try:
                    # Get or create company
                    company_name = record.get('company_name', 'Unknown Company')
                    company_nip = record.get('company_nip')
                    company_id = self.db.add_company(name=company_name, nip=company_nip)
                    
                    if not company_id:
                        skipped += 1
                        continue
                    
                    # Get or create order
                    order_id = self.db.add_order(
                        company_id=company_id,
                        date_issued=record.get('date_issued'),
                        document_number=record.get('document_number')
                    )
                    
                    if not order_id:
                        skipped += 1
                        continue
                    
                    # Add approved record
                    success = self.db.add_approved_record(
                        order_id=order_id,
                        date=record['date'],
                        filename=record['filename'],
                        filepath=record['filepath']
                    )
                    if success:
                        added += 1
                    else:
                        skipped += 1
                except Exception as e:
                    print(f"Error importing record: {e}")
                    skipped += 1
            
            return f"Baza danych połączona ({added} nowych, {skipped} pominięto)"
            
        except Exception as e:
            return f"Błąd łączenia baz: {str(e)}"
    
    def import_from_excel_file(self, excel_path: str) -> Tuple[bool, str]:
        """
        Import approved records from an ApprovedRecords.xlsx file.
        Reads the file paths from Excel and automatically copies Excel files
        and their associated images to the application.
        
        Args:
            excel_path: Path to ApprovedRecords.xlsx file or similar Excel with file paths
            
        Returns:
            Tuple of (success, message)
        """
        try:
            if not os.path.exists(excel_path):
                return False, "Plik nie istnieje"
            
            wb = load_workbook(excel_path)
            ws = wb['Approved']
            
            imported = 0
            skipped = 0
            errors = []
            files_copied = 0
            images_copied = 0
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 4:
                    continue
                
                date_value = row[0]
                company = row[1]
                filename = row[2]
                filepath = row[3]
                
                # Skip if filepath is empty
                if not filepath:
                    skipped += 1
                    continue
                
                # Normalize date
                if hasattr(date_value, 'strftime'):
                    date_str = date_value.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_value) if date_value else datetime.now().strftime('%Y-%m-%d')
                
                company_str = str(company).strip() if company else "Nieznana Firma"
                filename_str = str(filename) if filename else os.path.basename(str(filepath))
                filepath_str = str(filepath)
                
                try:
                    # Check if source Excel file exists
                    if not os.path.exists(filepath_str):
                        errors.append(f"Wiersz {row_idx}: Plik nie istnieje: {filepath_str}")
                        skipped += 1
                        continue
                    
                    # Determine destination in approved directory
                    dest_dir = os.path.join(APPROVED_DIRECTORY, company_str)
                    os.makedirs(dest_dir, exist_ok=True)
                    
                    dest_excel_path = os.path.join(dest_dir, filename_str)
                    
                    # Copy Excel file (skip if already exists with same content)
                    if not os.path.exists(dest_excel_path):
                        shutil.copy2(filepath_str, dest_excel_path)
                        files_copied += 1
                    
                    # Find and copy associated image
                    source_dir = os.path.dirname(filepath_str)
                    base_name = os.path.splitext(filename_str)[0]
                    
                    image_copied = False
                    for ext in ['.jpg', '.jpeg', '.png', '.bmp', '.JPG', '.JPEG', '.PNG', '.BMP']:
                        img_src = os.path.join(source_dir, base_name + ext)
                        if os.path.exists(img_src):
                            img_dest = os.path.join(dest_dir, base_name + os.path.splitext(img_src)[1])
                            if not os.path.exists(img_dest):
                                shutil.copy2(img_src, img_dest)
                                images_copied += 1
                            image_copied = True
                            break
                    
                    # Add to database with normalized schema
                    # 1. Get or create company
                    company_id = self.db.add_company(company_str)
                    
                    # 2. Create order
                    order_id = self.db.add_order(company_id, date_str)
                    
                    # 3. Add approved record to order
                    record_id = self.db.add_approved_record(
                        order_id=order_id,
                        date=date_str,
                        filename=filename_str,
                        filepath=dest_excel_path
                    )
                    
                    if record_id:
                        imported += 1
                    else:
                        skipped += 1
                        
                except Exception as e:
                    errors.append(f"Wiersz {row_idx}: {str(e)}")
                    skipped += 1
            
            wb.close()
            
            result_msg = f"✓ Zaimportowano do bazy: {imported}\n"
            result_msg += f"✓ Skopiowano plików Excel: {files_copied}\n"
            result_msg += f"✓ Skopiowano zdjęć: {images_copied}\n"
            result_msg += f"⊘ Pominięto (duplikaty/błędy): {skipped}"
            
            if errors:
                result_msg += f"\n\n⚠ Błędy ({len(errors)}):\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    result_msg += f"\n... i {len(errors) - 5} więcej"
            
            return True, result_msg
            
        except Exception as e:
            return False, f"Błąd importu z Excel: {str(e)}"
    
    def import_folder_batch(self, folder_path: str, status: str = "approved") -> Tuple[bool, str]:
        """
        Import a whole folder of Excel files with images.
        
        Args:
            folder_path: Path to folder containing Excel files and images
            status: "approved" or "unapproved"
            
        Returns:
            Tuple of (success, message)
        """
        try:
            if not os.path.exists(folder_path):
                return False, "Folder nie istnieje"
            
            excel_files = []
            
            # Find all Excel files in folder and subfolders
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    if file.endswith('.xlsx') and not file.startswith('~'):
                        excel_files.append(os.path.join(root, file))
            
            if not excel_files:
                return False, "Nie znaleziono plików Excel w folderze"
            
            imported = 0
            errors = []
            
            for excel_path in excel_files:
                try:
                    filename = os.path.basename(excel_path)
                    
                    # Parse date and company from filename or Excel content
                    wb = load_workbook(excel_path, data_only=True)
                    ws = wb.active
                    
                    date_value = ws['D1'].value
                    company_value = ws['B1'].value
                    
                    wb.close()
                    
                    # Format date
                    if hasattr(date_value, 'strftime'):
                        date_str = date_value.strftime('%Y-%m-%d')
                    elif isinstance(date_value, str):
                        date_str = date_value[:10]  # Assume yyyy-MM-dd format
                    else:
                        date_str = datetime.now().strftime('%Y-%m-%d')
                    
                    company_str = str(company_value).strip() if company_value else "Nieznana Firma"
                    
                    # Determine destination
                    if status == "approved":
                        dest_dir = os.path.join(APPROVED_DIRECTORY, company_str)
                        os.makedirs(dest_dir, exist_ok=True)
                        
                        dest_path = os.path.join(dest_dir, filename)
                        shutil.copy2(excel_path, dest_path)
                        
                        # Copy linked image
                        base_name = os.path.splitext(filename)[0]
                        for ext in ['.jpg', '.jpeg', '.png', '.bmp']:
                            img_src = os.path.join(os.path.dirname(excel_path), base_name + ext)
                            if os.path.exists(img_src):
                                img_dest = os.path.join(dest_dir, base_name + ext)
                                shutil.copy2(img_src, img_dest)
                                break
                        
                        # Add to database
                        self.db.add_approved_record(
                            date=date_str,
                            company=company_str,
                            filename=filename,
                            filepath=dest_path
                        )
                    else:
                        # Copy to unapproved folder
                        dest_path = os.path.join(REPORTS_ROOT, filename)
                        shutil.copy2(excel_path, dest_path)
                        
                        # Copy linked image
                        base_name = os.path.splitext(filename)[0]
                        for ext in ['.jpg', '.jpeg', '.png', '.bmp']:
                            img_src = os.path.join(os.path.dirname(excel_path), base_name + ext)
                            if os.path.exists(img_src):
                                img_dest = os.path.join(REPORTS_ROOT, base_name + ext)
                                shutil.copy2(img_src, img_dest)
                                break
                    
                    imported += 1
                    
                except Exception as e:
                    errors.append(f"{filename}: {str(e)}")
            
            result_msg = f"Zaimportowano: {imported} z {len(excel_files)} plików"
            if errors:
                result_msg += f"\n\nBłędy ({len(errors)}):\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    result_msg += f"\n... i {len(errors) - 5} więcej"
            
            return True, result_msg
            
        except Exception as e:
            return False, f"Błąd importu folderu: {str(e)}"
