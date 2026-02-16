import os
import re
import json
import shutil
import time
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import google.generativeai as genai
from google.api_core.exceptions import ServiceUnavailable
import config


class ImageTransformer:
    """Handles transformation of images to Excel reports using Gemini API."""
    
    def __init__(self, api_key=None):
        """Initialize the transformer with API key."""
        self.api_key = api_key or config.get_gemini_api_key()
        if not self.api_key:
            raise ValueError("GEMINI_API_KEY not found in environment or settings")
        genai.configure(api_key=self.api_key)
    
    def query_gemini_combined(self, image_path: str, model: str = "gemini-3-flash-preview") -> dict:
        """
        Send all 4 prompts in ONE API call (4x faster than separate calls).
        Returns dict with: odbiorca, nr_dokumentu, data_wystawienia, dane
        """
        image_file = genai.upload_file(path=image_path)

        combined_prompt = """Extract 4 pieces from this document:
1. 'ODBIORCA' cell (company): "ODBIORCA: [text]" (take the content of the whole cell, but exclude 'ODBIORCA' label)
2. 'Nr dokumentu' (document number): "Nr dokumentu: [text]"
3. 'Data wystawienia' (date, DD.MM.YYYY): "Data wystawienia: [text]"
4. All table data in pipe format: "|Lp|Nazwa|Ilość|Uwagi|Ilość|Stan poprzedni|Stan po wymianie|"

Use EXACTLY these formats, one item per line, then output table rows only."""

        models_to_try = [model, "gemini-2.5-flash", "gemini-2.5-pro", "gemini-3-flash-preview"]
        seen = set()
        unique_models = [m for m in models_to_try if not (m in seen or seen.add(m))]

        for model_name in unique_models:
            max_retries = 2
            wait_time = 1

            for attempt in range(max_retries):
                try:
                    model_obj = genai.GenerativeModel(model_name)
                    response = model_obj.generate_content([combined_prompt, image_file])

                    response_text = response.text
                    lines = response_text.strip().split('\n')
                    result = {
                        'odbiorca': 'UNKNOWN',
                        'nr_dokumentu': 'UNKNOWN',
                        'data_wystawienia': 'UNKNOWN',
                        'dane': ''
                    }
                    
                    table_lines = []
                    for line in lines:
                        line = line.strip()
                        if line.startswith('ODBIORCA:'):
                            result['odbiorca'] = self.extract_after_colon(line)
                        elif line.startswith('Nr dokumentu:'):
                            result['nr_dokumentu'] = self.extract_after_colon(line)
                        elif line.startswith('Data wystawienia:'):
                            result['data_wystawienia'] = self.extract_after_colon(line)
                        elif '|' in line:
                            table_lines.append(line)
                    
                    result['dane'] = '\n'.join(table_lines)
                    return result

                except ServiceUnavailable as e:
                    if attempt < max_retries - 1:
                        print(f"Model {model_name} 503. Retrying in {wait_time}s...")
                        time.sleep(wait_time)
                        wait_time *= 2
                    else:
                        print(f"Model {model_name} unavailable")
                        break
                except Exception as e:
                    print(f"Model {model_name} failed: {e}")
                    break

        return {'odbiorca': 'UNKNOWN', 'nr_dokumentu': 'UNKNOWN', 'data_wystawienia': 'UNKNOWN', 'dane': '', 'error': 'All models failed'}
    
    def query_gemini_with_image(self, prompt: str, image_path: str, model: str = "gemini-3-flash-preview") -> str:
        """
        Send a prompt and image to a Google Gemini multimodal model and return the full response text.
        Implements exponential backoff retry logic for 503 errors and falls back to alternate models.
        """
        # Upload image file once; reuse across model attempts
        image_file = genai.upload_file(path=image_path)

        # Models to try in order (avoid duplicates while preserving order)
        primary_and_fallbacks = [model, "gemini-3-flash-preview", "gemini-2.5-flash", "gemini-2.5-pro"]
        seen = set()
        models_to_try = []
        for m in primary_and_fallbacks:
            if m not in seen:
                models_to_try.append(m)
                seen.add(m)

        for model_name in models_to_try:
            max_retries = 5
            wait_time = 1  # Start with 1 second

            for attempt in range(max_retries):
                try:
                    model_obj = genai.GenerativeModel(model_name)
                    response = model_obj.generate_content([prompt, image_file])

                    try:
                        return response.text
                    except Exception as e:
                        return json.dumps({"error": str(e)}, indent=2)

                except ServiceUnavailable as e:
                    if attempt < max_retries - 1:
                        print(
                            f"Model {model_name} returned 503. Retrying in {wait_time}s... "
                            f"(Attempt {attempt + 1}/{max_retries})"
                        )
                        time.sleep(wait_time)
                        wait_time *= 2
                    else:
                        print(f"Model {model_name} unavailable after {max_retries} attempts: {e}")
                        break  # Try next model

                except Exception as e:
                    print(f"Model {model_name} failed with error: {e}. Trying next model...")
                    break  # Try next model

        # All models failed
        return json.dumps({"error": "All models failed (gemini-3-flash-preview, gemini-2.5-flash, gemini-2.5-pro)"}, indent=2)

    def extract_after_colon(self, text: str) -> str:
        """
        Return the substring after the first ':' in `text`, stripped of surrounding whitespace.
        """
        if not isinstance(text, str):
            return text
        idx = text.find(":")
        if idx == -1:
            s = text.strip()
        else:
            s = text[idx + 1:].strip()

        # Remove a trailing dot if present
        while s.endswith('.'):
            s = s[:-1].rstrip()

        return s

    def normalize_invoice_number(self, nr_dokumentu):
        if not isinstance(nr_dokumentu, str):
            return nr_dokumentu
        trimmed = nr_dokumentu.strip()
        if len(trimmed) >= 3 and trimmed[-3:].upper() == "FUS":
            return trimmed[:-3] + "FVS"
        return trimmed

    def parse_date_flexible(self, date_text: str) -> datetime:
        """
        Parse `date_text` using several common formats and return a datetime.
        Handles dates with or without time information.
        """
        if not isinstance(date_text, str):
            raise ValueError("date value must be a string")
        s = date_text.strip()

        # Remove time portion if present (e.g., "19.01.2026 10:59:27" -> "19.01.2026")
        s = s.split()[0]

        # Try common explicit formats
        formats = (
            "%d.%m.%Y",
            "%d.%m.%y",  # DD.MM.YY format (e.g., 03.01.26)
            "%d/%m/%Y",
            "%d/%m/%y",
            "%d-%m-%Y",
            "%d-%m-%y",
            "%Y-%m-%d",
            "%Y.%m.%d",
            "%d %m %Y",
        )
        for fmt in formats:
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue

        # Try normalizing separators
        s_normalized = re.sub(r"[^0-9]", ".", s)
        try:
            return datetime.strptime(s_normalized, "%d.%m.%Y")
        except ValueError:
            pass

        raise ValueError(f"Unrecognized date format: {date_text!r}")

    def process_image_file(self, image_path: str, base_folder: str = "Reports") -> str:
        """
        Run Gemini extraction for a single image and write the Excel report.
        Uses combined API call for 4x faster processing.
        Returns the written Excel path.
        """
        if not os.path.isfile(image_path):
            raise FileNotFoundError(f"Image not found: {image_path}")

        errors = []

        # Use combined API call (1 call instead of 4 = 4x faster!)
        try:
            result = self.query_gemini_combined(image_path)
            odbiorca = result.get('odbiorca', 'UNKNOWN')
            nr_dokumentu = result.get('nr_dokumentu', 'UNKNOWN')
            data_wystawienia = result.get('data_wystawienia', 'UNKNOWN')
            dane = result.get('dane', '')
            
            if result.get('error'):
                errors.append(f"API response: {result['error']}")
        except Exception as e:
            errors.append(f"API query error: {e}")
            odbiorca = "UNKNOWN"
            nr_dokumentu = "UNKNOWN"
            data_wystawienia = "UNKNOWN"
            dane = ""

        nr_dokumentu = self.normalize_invoice_number(nr_dokumentu)

        # Parse date with fallback to today
        try:
            date_obj = self.parse_date_flexible(data_wystawienia)
        except Exception as e:
            errors.append(f"Date parse error: {e}")
            date_obj = datetime.today()
        date_str = date_obj.strftime("%Y-%m-%d")

        safe_company = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in odbiorca)
        if not safe_company:
            safe_company = "UNKNOWN"
        
        # Truncate to 30 characters max and strip whitespace to avoid long file paths and Windows filename limits
        safe_company = safe_company[:30].strip()
        if not safe_company:
            safe_company = "UNKNOWN"

        # Use os.path.normpath to ensure proper path construction
        company_folder = os.path.normpath(os.path.join(base_folder, safe_company))
        os.makedirs(company_folder, exist_ok=True)
        
        # Generate unique filename if file already exists
        base_filename = f"{date_str}_{safe_company}"
        file_name = f"{base_filename}.xlsx"
        file_path = os.path.normpath(os.path.join(company_folder, file_name))
        
        # If file exists, add a counter suffix to make it unique
        counter = 1
        while os.path.exists(file_path):
            file_name = f"{base_filename}_{counter}.xlsx"
            file_path = os.path.normpath(os.path.join(company_folder, file_name))
            counter += 1

        # Build table even if parsing failed
        table_block = re.findall(r"\|.*\|", dane)
        table_block = [row for row in table_block if "---" not in row]
        rows = [[cell.strip() for cell in row.split("|")[1:-1]] for row in table_block]

        expected_header = [
            "Lp",
            "Nazwa",
            "Ilość",
            "Uwagi",
            "Ilość",
            "Stan poprzedni",
            "Stan po wymianie",
        ]

        def looks_like_header(cells):
            # Basic heuristic: contains key labels and matches expected column count
            if len(cells) != len(expected_header):
                return False
            lowered = [c.lower() for c in cells]
            return any("lp" in c for c in lowered) and any("nazwa" in c for c in lowered)

        if not rows:
            # Fallback table with raw response or error info
            header = ["Notice", "Value"]
            data = [
                ["Table parsing", "No valid table found"],
                ["API response", dane[:500] if dane else "<empty>"]
            ]
        elif looks_like_header(rows[0]):
            header = rows[0]
            data = rows[1:]
        else:
            # Header missing or malformed: inject expected header and keep all rows as data
            header = expected_header
            data = rows

        # Ensure all data rows match header column count
        expected_cols = len(header)
        normalized_data = []
        for row in data:
            if len(row) > expected_cols:
                # Truncate extra columns and convert to strings
                processed_row = [str(cell) if cell is not None else "" for cell in row[:expected_cols]]
            elif len(row) < expected_cols:
                # Pad with empty strings and convert all cells to strings
                padded_row = row + [""] * (expected_cols - len(row))
                processed_row = [str(cell) if cell is not None else "" for cell in padded_row]
            else:
                # Convert all cells to strings
                processed_row = [str(cell) if cell is not None else "" for cell in row]
            
            # Replace "nan" with empty strings
            processed_row = [
                "" if cell.lower() == "nan" else cell
                for cell in processed_row
            ]
            normalized_data.append(processed_row)

        df_table = pd.DataFrame(normalized_data, columns=header)

        # Meta sheet with possible errors
        # Ensure all metadata values are strings to prevent NaN in output
        meta_row = [
            "Odbiorca",
            str(odbiorca) if odbiorca else "",
            "Data wystawienia",
            str(data_wystawienia) if data_wystawienia else "",
            "Nr dokumentu",
            str(nr_dokumentu) if nr_dokumentu else ""
        ]
        
        df_meta = pd.DataFrame([meta_row])
        if errors:
            # Add an Errors row for visibility
            df_meta = pd.concat([df_meta, pd.DataFrame([["Errors", " | ".join(errors)]])], ignore_index=True)

        # Replace NaN with empty strings before writing to Excel
        df_meta = df_meta.fillna('').astype(str)
        df_table = df_table.fillna('').astype(str)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_meta.to_excel(writer, sheet_name="Sheet1", index=False, header=False, startrow=0)
            df_table.to_excel(writer, sheet_name="Sheet1", index=False, startrow=2)

        wb = load_workbook(file_path)
        ws = wb["Sheet1"]

        for col in [1, 3, 5]:
            ws.cell(row=1, column=col).font = Font(bold=True)

        for col in [2, 4, 6]:
            ws.cell(row=1, column=col).alignment = Alignment(wrap_text=True)

        ws.row_dimensions[1].height = 80
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['F'].width = 15

        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        def to_num(x):
            if x is None or (isinstance(x, str) and x.strip() == ""):
                return None
            try:
                return float(str(x).replace(",", "."))
            except:
                return None

        highlighted = 0

        for row_idx in range(4, ws.max_row + 1):
            c = to_num(ws.cell(row=row_idx, column=3).value)
            e = to_num(ws.cell(row=row_idx, column=5).value)
            f = to_num(ws.cell(row=row_idx, column=6).value)
            g = to_num(ws.cell(row=row_idx, column=7).value)

            # Treat None as 0.0 for math, but keep None if column is completely empty
            safe_f = f if f is not None else 0.0
            expected = None
            
            if c is None and e is None:
                expected = f
            elif c is not None and e is None:
                expected = safe_f + c
            elif c is not None and e is not None:
                # Logic: Previous + Delivery - Return (validates even when C == E)
                expected = safe_f + c - e
            elif c is None and e is not None:
                expected = safe_f - e

            # Compare with rounding to prevent floating point errors
            if expected is not None and g is not None:
                if round(float(g), 2) != round(float(expected), 2):
                    ws.cell(row=row_idx, column=7).fill = red_fill
                    highlighted += 1

        wb.save(file_path)
        
        # Copy source image to same folder with matching name (extract base name from file_name)
        image_extension = os.path.splitext(image_path)[1]
        # Get the xlsx filename without extension to use as base for image name
        excel_base = os.path.splitext(file_name)[0]
        image_copy_name = f"{excel_base}{image_extension}"
        image_copy_path = os.path.normpath(os.path.join(company_folder, image_copy_name))
        shutil.copy2(image_path, image_copy_path)
        
        print(f"✓ Saved Excel: {file_name}")
        print(f"✓ Saved Image: {image_copy_name}")
        
        return file_path, highlighted

    def process_multiple_images(self, image_paths: list, base_folder: str = "Reports"):
        """
        Process multiple images and return results.
        Returns a list of tuples: (image_path, success, result_or_error)
        """
        results = []
        for img_path in image_paths:
            try:
                output_path, highlighted = self.process_image_file(img_path, base_folder)
                results.append((img_path, True, output_path, highlighted))
            except Exception as e:
                results.append((img_path, False, str(e), 0))
        return results
