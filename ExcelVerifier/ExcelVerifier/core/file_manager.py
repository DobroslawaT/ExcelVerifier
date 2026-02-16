import os
from openpyxl import load_workbook
# Make sure your config.py actually defines these variables
from config import REPORTS_ROOT, APPROVED_FILE, DATABASE_FILE

class FileManager:
    def get_unapproved_reports(self):
        """
        Recursively scans REPORTS_ROOT for Excel files and returns a list 
        of absolute paths for files that are NOT listed in APPROVED_FILE.
        """
        approved_names = self._get_approved_filenames()
        print(f"DEBUG: Approved filenames: {approved_names}")
        
        unapproved_paths = []
        all_excel_files = []

        # Walk the reports folder and collect .xlsx/.xlsm/.xls files
        if os.path.exists(REPORTS_ROOT):
            for dirpath, dirnames, filenames in os.walk(REPORTS_ROOT):
                for fname in filenames:
                    # Skip temporary Excel lock files
                    if fname.startswith('~$'):
                        print(f"DEBUG: Skipping temp lock file: {fname}")
                        continue
                    
                    # Check extension
                    low = fname.lower()
                    if low.endswith(('.xlsx', '.xlsm', '.xls')):
                        all_excel_files.append(fname)
                        full_path = os.path.join(dirpath, fname)
                        
                        # Check if already approved
                        if fname in approved_names:
                            print(f"DEBUG: Skipping approved file: {fname}")
                            continue
                        
                        unapproved_paths.append(full_path)
                        print(f"DEBUG: Found unapproved report: {full_path}")

        # Sort results for consistent order (A-Z)
        unapproved_paths.sort()
        print(f"DEBUG: Total Excel files found: {len(all_excel_files)}")
        print(f"DEBUG: Total approved files: {len(approved_names)}")
        print(f"DEBUG: Total unapproved reports: {len(unapproved_paths)}")
        return unapproved_paths

    def _get_approved_filenames(self):
        """
        Helper: Gets approved filenames from the database (SQLite).
        Returns a set of filenames that have already been approved.
        Falls back to legacy Excel file if database is unavailable.
        """
        approved_names = set()
        
        try:
            # Try to load from database
            from core.database_handler import DatabaseHandler
            db = DatabaseHandler(DATABASE_FILE)
            approved_records = db.get_all_approved_records()
            
            for record in approved_records:
                filename = record.get('filename')
                if filename:
                    approved_names.add(str(filename))
                    
            print(f"DEBUG: Loaded {len(approved_names)} approved filenames from database")
            return approved_names
            
        except Exception as e:
            print(f"Warning: Could not read approved records from database: {e}")
            # Fall through to legacy Excel file
        
        # Fallback: Try legacy Excel file
        try:
            if not os.path.exists(APPROVED_FILE):
                return approved_names

            wb = load_workbook(APPROVED_FILE, read_only=True)
            
            if 'Approved' in wb.sheetnames:
                ws = wb['Approved']
                
                # 1. Try to find the 'Filename' column dynamically
                fname_col_idx = None
                
                # Read headers (first row)
                headers = []
                # ws.iter_rows is more efficient in read_only mode
                rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
                for row in rows:
                    headers = [str(cell) if cell is not None else '' for cell in row]
                    break
                
                if 'Filename' in headers:
                    # headers index is 0-based, but we need the value from the row
                    fname_col_idx = headers.index('Filename')

                # 2. Iterate through data rows
                row_iter = ws.iter_rows(min_row=2, values_only=True)
                for row in row_iter:
                    val = None
                    if fname_col_idx is not None and fname_col_idx < len(row):
                        val = row[fname_col_idx]
                    elif len(row) >= 3:
                        # Fallback: assume column 3 (index 2) if header missing
                        val = row[2]
                    
                    if val:
                        approved_names.add(str(val))
                        
            wb.close()
            
        except Exception as e2:
            print(f"Warning: Could not read legacy approved records: {e2}")
            
        return approved_names
    
    def get_approved_reports(self):
        """
        Returns a list of absolute paths for all approved reports.
        Reads from APPROVED_FILE and finds the actual files in REPORTS_ROOT.
        """
        approved_names = self._get_approved_filenames()
        approved_paths = []
        
        # Walk the reports folder and collect approved files
        if os.path.exists(REPORTS_ROOT):
            for dirpath, dirnames, filenames in os.walk(REPORTS_ROOT):
                for fname in filenames:
                    # Skip temporary Excel lock files
                    if fname.startswith('~$'):
                        continue
                    
                    # Check extension
                    low = fname.lower()
                    if low.endswith(('.xlsx', '.xlsm', '.xls')):
                        # Only include if it's in approved list
                        if fname in approved_names:
                            full_path = os.path.join(dirpath, fname)
                            approved_paths.append(full_path)
        
        # Sort results for consistent order
        approved_paths.sort()
        return approved_paths