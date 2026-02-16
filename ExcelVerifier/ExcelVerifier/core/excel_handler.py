import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from config import APPROVED_FILE, REPORTING_DATA_FILE, COMPANY_DB_FILE, DATABASE_FILE, APPROVED_DIRECTORY
from core.image_transformer import ImageTransformer
from core.company_db import load_company_db, normalize_nip
from core.database_handler import DatabaseHandler

class ExcelHandler:
    def __init__(self):
        self.current_workbook = None
        self.current_df = None
        self.file_path = None
        self.original_fills = {} # To store original formatting before validation
        self.db = DatabaseHandler(DATABASE_FILE)

    def load_file(self, file_path):
        """Loads excel and returns a DataFrame for the UI to display."""
        self.file_path = file_path
        self.current_workbook = load_workbook(file_path)
        ws = self.current_workbook.active
        
        # Load data into Pandas
        # Note: We treat the sheet as raw data (no headers) to match your original logic
        # where row 1 was editable.
        self.current_df = pd.DataFrame(ws.values)
        return self.current_df

    def save_data(self, ui_table_data):
        """
        Receives raw data from UI, updates workbook, 
        runs the math validation logic (red coloring),
        and saves the file.
        Detects if 'Odbiorca' (row 1, col 2) changed and reorganizes files accordingly.
        """
        if not self.current_workbook:
            return

        ws = self.current_workbook.active
        
        # Helper to safely copy a fill (Fixes 'StyleProxy' error)
        def clone_fill(proxy):
            if not proxy: return None
            return PatternFill(
                start_color=proxy.start_color,
                end_color=proxy.end_color,
                fill_type=proxy.fill_type
            )

        # Capture old Odbiorca value before update
        old_odbiorca = ws.cell(row=1, column=2).value or "UNKNOWN"

        # 1. Capture original fills for Column G (7)
        # We explicitly create a NEW PatternFill object here.
        self.original_fills = {}
        for row_idx in range(4, ws.max_row + 1):
            source_fill = ws.cell(row=row_idx, column=7).fill
            self.original_fills[row_idx] = clone_fill(source_fill)

        # 2. Update Workbook with data from UI
        for i, row_data in enumerate(ui_table_data):
            for j, text_val in enumerate(row_data):
                cell = ws.cell(row=i+1, column=j+1)
                new_val = self._convert_type(text_val, cell.value)
                cell.value = new_val

        # Check if Odbiorca changed
        new_odbiorca = ws.cell(row=1, column=2).value or "UNKNOWN"
        odbiorca_changed = str(old_odbiorca).strip() != str(new_odbiorca).strip()

        # 3. Run Validation Logic (Red Coloring)
        self._apply_validation_coloring(ws)

        # 4. Save the main file
        try:
            self.current_workbook.save(self.file_path)
        except PermissionError:
            raise Exception(f"Permission denied: Close '{os.path.basename(self.file_path)}' in Excel and try again.")

        # 5. Reorganize files if Odbiorca changed
        if odbiorca_changed and self.file_path:
            try:
                transformer = ImageTransformer()
                new_excel_path, new_image_path = transformer.reorganize_files_by_company(
                    self.file_path, 
                    new_odbiorca,
                    base_folder="Reports"
                )
                # Update the file path to point to the new location
                self.file_path = new_excel_path
                self.current_workbook = load_workbook(new_excel_path)
                print(f"✓ Files reorganized for '{new_odbiorca}'")
            except Exception as e:
                print(f"⚠ Warning: Could not reorganize files: {e}")
                # Continue anyway; file was saved to old location

        # 6. Sync with reportingData.xlsx
        self._update_reporting_data()

    def get_formatting(self):
        """
        Returns a dictionary mapping coordinates to colors:
        { (row_idx, col_idx): {'bg': '#FF0000', 'fg': '#000000'} }
        """
        if not self.current_workbook:
            return {}

        # Import helper here to prevent circular import issues
        from ui.utils import resolve_excel_color 
        
        ws = self.current_workbook.active
        styles = {}

        # Scan all cells that have data or formatting
        for row in ws.iter_rows():
            for cell in row:
                # Convert 1-based Excel index to 0-based UI index
                r = cell.row - 1
                c = cell.column - 1
                
                # Resolve colors
                bg_hex = resolve_excel_color(cell.fill.fgColor)
                fg_hex = resolve_excel_color(cell.font.color)
                
                # Get font styles
                is_bold = cell.font.bold if cell.font.bold is not None else False
                is_italic = cell.font.italic if cell.font.italic is not None else False
                is_underline = cell.font.underline is not None and cell.font.underline != 'none'

                if bg_hex or fg_hex or is_bold or is_italic or is_underline:
                    styles[(r, c)] = {
                        'bg': bg_hex, 
                        'fg': fg_hex,
                        'bold': is_bold,
                        'italic': is_italic,
                        'underline': is_underline
                    }
        
        return styles

    def approve_report(self, filename, date_part, company_part, full_path):
        """
        1. Move file from Niezatwierdzone to Zatwierdzone
        2. Move linked image (if exists) to Zatwierdzone
        3. Write metadata to database with new path
        4. Append detailed rows to database
        """
        import shutil
        import os
        
        # Determine the destination directory (Zatwierdzone)
        approved_dir = APPROVED_DIRECTORY
        if not approved_dir:
            raise Exception("APPROVED_DIRECTORY is not configured")
        
        os.makedirs(approved_dir, exist_ok=True)
        
        # Create the destination file path
        approved_path = os.path.join(approved_dir, filename)
        
        # Move the Excel file from Niezatwierdzone to Zatwierdzone
        try:
            print(f"[APPROVE] Moving Excel from: {full_path}")
            print(f"[APPROVE] Moving Excel to: {approved_path}")
            shutil.move(full_path, approved_path)
            print(f"[APPROVE] Excel file moved successfully")
        except Exception as e:
            print(f"[APPROVE] Error moving Excel file: {e}")
            raise Exception(f"Nie udało się przenieść pliku do zatwierdzonego folderu: {e}")
        
        # Move the linked image if it exists
        try:
            excel_dir = os.path.dirname(full_path)
            excel_base = os.path.splitext(filename)[0]
            
            image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff']
            for ext in image_extensions:
                source_image = os.path.join(excel_dir, excel_base + ext)
                if os.path.exists(source_image):
                    dest_image = os.path.join(approved_dir, excel_base + ext)
                    print(f"[APPROVE] Moving image from: {source_image}")
                    print(f"[APPROVE] Moving image to: {dest_image}")
                    shutil.move(source_image, dest_image)
                    print(f"[APPROVE] Image moved successfully")
                    break
        except Exception as e:
            print(f"[APPROVE] Warning: Could not move image: {e}")
            # Don't fail the approval if image move fails
        
        # A. Update database with the new approved path
        self._write_approval_metadata(filename, date_part, company_part, approved_path)
        
        # B. Append details to database
        self._append_detailed_records()
        self._write_approval_metadata(filename, date_part, company_part, approved_path)
        
        # B. Append details to database
        self._append_detailed_records()

    # =========================================
    # INTERNAL HELPER METHODS
    # =========================================

    def _convert_type(self, text, original_value):
        """Helps preserve numbers/bools when saving string data from UI."""
        if text == "" or text == "-":
            return None
            
        if isinstance(original_value, bool):
            return text.lower() == "true"
        
        # Try converting to number if it looks like one
        try:
            # Check if it was an int originally
            if isinstance(original_value, int):
                return int(float(text))
            # Check if it was a float originally
            if isinstance(original_value, float):
                return float(text)
            
            # If original was None or String, but new text is numeric, convert it
            if text.replace('.', '', 1).replace('-', '', 1).isdigit():
                if '.' in text:
                    return float(text)
                return int(text)
        except:
            pass
            
        return text

    def _normalize_invoice_number(self, nr_dok):
        if not isinstance(nr_dok, str):
            return nr_dok
        trimmed = nr_dok.strip()
        if len(trimmed) >= 3 and trimmed[-3:].upper() == "FUS":
            return trimmed[:-3] + "FVS"
        return trimmed

    def _to_num(self, x):
        """
        User-provided helper: robustly converts string/mix to float.
        Handles commas and empty strings.
        """
        if x is None or (isinstance(x, str) and x.strip() == ""):
            return None
        # Handle em dash (—) and hyphen (-) as None
        if isinstance(x, str) and x.strip() in ("—", "-", "–"):
            return None
        try:
            # Replace comma with dot for Polish/European decimals
            return float(str(x).replace(",", "."))
        except:  
            return None

    def extract_nip(self, text):
        """
        Extracts NIP from text. Searches for patterns like:
        - 10 digits: 1234567890
        - Formatted: XXX-XX-XX-XXX (like 123-45-67-890)
        - Formatted: XXX-XXX-XX-XX (like 123-456-78-90)
        
        Returns the NIP as 10 digits without dashes (e.g., 1234567890) or None if not found.
        """
        if not text:
            return None
        
        text = str(text).strip()
        import re
        
        # Try formatted pattern: XXX-XX-XX-XXX and remove dashes
        formatted_match = re.search(r'(\d{3})-(\d{2})-(\d{2})-(\d{3})', text)
        if formatted_match:
            return f"{formatted_match.group(1)}{formatted_match.group(2)}{formatted_match.group(3)}{formatted_match.group(4)}"
        
        # Try formatted pattern: XXX-XXX-XX-XX and remove dashes
        formatted_match2 = re.search(r'(\d{3})-(\d{3})-(\d{2})-(\d{2})', text)
        if formatted_match2:
            return f"{formatted_match2.group(1)}{formatted_match2.group(2)}{formatted_match2.group(3)}{formatted_match2.group(4)}"
        
        # Try 10 consecutive digits (must be isolated - not part of a longer sequence)
        digits_match = re.search(r'(?:^|\s|[^\d])(\d{10})(?:\s|$|[^\d])', text)
        if digits_match:
            return digits_match.group(1)
        
        return None

    def _normalize_company_name(self, value):
        return " ".join(str(value).strip().lower().split())

    def _fill_missing_nip_from_db(self, df):
        if df is None or 'Odbiorca' not in df.columns or 'NIP' not in df.columns:
            return df

        companies = load_company_db(COMPANY_DB_FILE)
        nip_map = {}
        for item in companies:
            name = item.get('name')
            nip = normalize_nip(item.get('nip'))
            if name and nip:
                nip_map[self._normalize_company_name(name)] = nip

        if not nip_map:
            return df

        def resolve_nip(row):
            existing_raw = row.get('NIP', '')
            if normalize_nip(existing_raw):
                return existing_raw
            key = self._normalize_company_name(row.get('Odbiorca', ''))
            return nip_map.get(key, existing_raw)

        df['NIP'] = df.apply(resolve_nip, axis=1)
        return df

    def _apply_validation_coloring(self, ws):
        """
        Runs the math checks (Columns C, E, F vs G).
        - If Wrong: Applies Red (FF0000).
        - If Correct: Clears Red (restores original color ONLY if it wasn't Red).
        """
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        clear_fill = PatternFill(fill_type=None)
        
        highlighted_count = 0
        print(f"--- Validating rows 4 to {ws.max_row} ---")

        for row_idx in range(4, ws.max_row + 1):
            # 1. Get Values
            c = self._to_num(ws.cell(row=row_idx, column=3).value)  # Col C (Dostawa)
            e = self._to_num(ws.cell(row=row_idx, column=5).value)  # Col E (Zwrot)
            f = self._to_num(ws.cell(row=row_idx, column=6).value)  # Col F (Stan Prev)
            g = self._to_num(ws.cell(row=row_idx, column=7).value)  # Col G (Stan Act)

            # 2. Calculate Expected
            # Treat None as 0.0 for math, but keep None if column is completely empty
            safe_f = f if f is not None else 0.0
            expected = None

            if c is None and e is None:
                expected = f 
            elif c is not None and e is None:
                expected = safe_f + c
            elif c is not None and e is not None:
                # Logic: Previous + Delivery - Return
                expected = safe_f + c - e
            elif c is None and e is not None:
                expected = safe_f - e

            # 3. Compare (Round to 2 decimals to prevent 10.0 != 10.00001 errors)
            is_error = False
            if expected is not None and g is not None:
                if round(float(g), 2) != round(float(expected), 2):
                    is_error = True
                    print(f"Row {row_idx} Error: Exp {expected} != Act {g}")

            # 4. Apply Colors
            cell_g = ws.cell(row=row_idx, column=7)

            if is_error:
                cell_g.fill = red_fill
                highlighted_count += 1
            else:
                # --- LOGIC TO CLEAR COLOR WHEN CORRECT ---
                # We want to restore the original background (e.g., gray header styles),
                # BUT NOT if the original background was the "Error Red" from a previous save.
                
                orig = self.original_fills.get(row_idx)
                was_red_originally = False

                # Check if the original was red (FF9999 or FF0000)
                if orig and orig.start_color and hasattr(orig.start_color, 'rgb'):
                    hex_val = str(orig.start_color.rgb).upper()
                    if 'FF9999' in hex_val or 'FF0000' in hex_val:
                        was_red_originally = True

                if orig and orig.fill_type and not was_red_originally:
                    # It was a nice formatting color (not red), restore it
                    cell_g.fill = orig
                else:
                    # It was red (or empty) before -> make it transparent now
                    cell_g.fill = clear_fill

        print(f"Validation finished. Errors marked: {highlighted_count}")

    def _update_reporting_data(self):
        """Updates rows in reportingData.xlsx based on matching Document #."""
        if not os.path.exists(self.file_path): return

        # Load current open sheet data
        ws_curr = self.current_workbook.active
        nr_dok = self._normalize_invoice_number(ws_curr['F1'].value)
        
        # Prepare list of rows to update
        report_rows = []
        for r in range(4, ws_curr.max_row + 1):
            nazwa = ws_curr.cell(row=r, column=2).value
            if not nazwa: continue
            # Collect row data tuple
            report_rows.append({
                'nr_dok': nr_dok,
                'nazwa': str(nazwa),
                'data_wyst': ws_curr['D1'].value,
                'odbiorca': ws_curr['B1'].value,
                'il_dost': ws_curr.cell(row=r, column=3).value,
                'il_zwrot': ws_curr.cell(row=r, column=5).value,
                'stan_prev': ws_curr.cell(row=r, column=6).value,
                'stan_po': ws_curr.cell(row=r, column=7).value
            })

        if not report_rows: return

        # Open Reporting DB
        target = REPORTING_DATA_FILE
        if not os.path.exists(target):
            wb = Workbook()
            ws = wb.active
            ws.title = 'Records'
            ws.append(['data wystawienia', 'Odbiorca', 'nr dokumentu', 'nazwa', 
                       'ilość dostawa', 'ilość zwrot', 'stan poprzedni', 'stan po wymianie'])
            wb.save(target)

        try:
            wb_db = load_workbook(target)
            if 'Records' not in wb_db.sheetnames:
                ws_db = wb_db.create_sheet('Records')
            else:
                ws_db = wb_db['Records']

            # Find columns
            headers = [c.value for c in ws_db[1]]
            try:
                col_map = {h: i+1 for i, h in enumerate(headers)}
                col_nr = col_map['nr dokumentu']
                col_nazwa = col_map['nazwa']
            except:
                return # Headers missing

            # Update loop
            for row_data in report_rows:
                match_found = False
                for db_r in range(2, ws_db.max_row + 1):
                    # Check match on Doc ID and Product Name
                    if (ws_db.cell(db_r, col_nr).value == row_data['nr_dok'] and 
                        str(ws_db.cell(db_r, col_nazwa).value) == row_data['nazwa']):
                        
                        # Update fields
                        ws_db.cell(db_r, col_map['data wystawienia'], row_data['data_wyst'])
                        ws_db.cell(db_r, col_map['Odbiorca'], row_data['odbiorca'])
                        ws_db.cell(db_r, col_map['ilość dostawa'], row_data['il_dost'])
                        ws_db.cell(db_r, col_map['ilość zwrot'], row_data['il_zwrot'])
                        ws_db.cell(db_r, col_map['stan poprzedni'], row_data['stan_prev'])
                        ws_db.cell(db_r, col_map['stan po wymianie'], row_data['stan_po'])
                        match_found = True
                        break
                
                # If no match found in DB, we DO NOT append here. 
                # Appending happens only on "Approve". 
                # This function only syncs edits to existing records.

            wb_db.save(target)
        except Exception as e:
            print(f"Failed to sync reporting data: {e}")

    def _write_approval_metadata(self, filename, date_part, company_part, full_path):
        """Writes approval record to database using new schema"""
        try:
            # Get or create company
            company_id = self.db.add_company(name=company_part, nip=None)
            if not company_id:
                raise Exception(f"Failed to create/get company: {company_part}")
            
            # Get document number and date from current workbook
            ws = self.current_workbook.active
            document_number = self._normalize_invoice_number(ws['F1'].value) if ws['F1'].value else None
            date_issued = ws['D1'].value if ws['D1'].value else date_part
            
            # Create order
            order_id = self.db.add_order(
                company_id=company_id,
                date_issued=date_issued,
                document_number=document_number
            )
            if not order_id:
                raise Exception("Failed to create order")
            
            # Create approved_record linked to order
            record_id = self.db.add_approved_record(
                order_id=order_id,
                date=date_part,
                filename=filename,
                filepath=full_path
            )
            if not record_id:
                print(f"Note: Record already exists for {filename}")
                return
            
            print(f"✓ Created approved record (Order #{order_id}, Record #{record_id})")
            
        except Exception as e:
            raise Exception(f"Failed to write approval metadata: {e}")
    
    def _append_detailed_records(self):
        """Appends ALL rows from current file to database (Used on Approval)"""
        from datetime import datetime
        
        ws_curr = self.current_workbook.active
        
        nr_dok = self._normalize_invoice_number(ws_curr['F1'].value)
        data_wyst = ws_curr['D1'].value
        odbiorca = ws_curr['B1'].value
        filename = os.path.basename(self.file_path) if self.file_path else ""
        
        # Normalize date to yyyy-MM-dd
        if isinstance(data_wyst, str):
            try:
                parsed = datetime.strptime(data_wyst.strip(), '%d.%m.%Y')
                data_wyst = parsed.strftime('%Y-%m-%d')
            except:
                try:
                    parsed = datetime.strptime(data_wyst.strip(), '%d.%m.%y')
                    data_wyst = parsed.strftime('%Y-%m-%d')
                except:
                    data_wyst = str(data_wyst)
        elif hasattr(data_wyst, 'strftime'):
            data_wyst = data_wyst.strftime('%Y-%m-%d')
        else:
            data_wyst = str(data_wyst)

        # Get the approved_record for this filename to get order_id
        approved_rec = self.db.get_approved_record(filename)
        if not approved_rec:
            raise Exception(f"Cannot find approved record for {filename}. Please approve the file first.")
        
        order_id = approved_rec['order_id']
        order_items = []
        
        for r in range(4, ws_curr.max_row + 1):
            nazwa = ws_curr.cell(row=r, column=2).value
            # Skip empty rows
            if not nazwa and not ws_curr.cell(row=r, column=3).value: 
                continue
            
            dost = ws_curr.cell(row=r, column=3).value
            zwrot = ws_curr.cell(row=r, column=5).value
            prev = ws_curr.cell(row=r, column=6).value
            po = ws_curr.cell(row=r, column=7).value
            
            # Get or create product
            product_name = str(nazwa).strip() if nazwa else ""
            if not product_name:
                continue
            
            product_id = self.db.add_product(name=product_name, code=None)
            if not product_id:
                print(f"  ⚠ Warning: Failed to create product '{product_name}'")
                continue
            
            order_items.append({
                'order_id': order_id,
                'product_id': product_id,
                'quantity_delivery': float(dost) if dost else 0.0,
                'quantity_return': float(zwrot) if zwrot else 0.0,
                'previous_state': float(prev) if prev else 0.0,
                'state_after': float(po) if po else 0.0
            })

        if not order_items:
            return

        # Insert into database
        try:
            count = self.db.add_order_items(order_items)
            print(f"✓ Added {count} order items to database")
        except Exception as e:
            raise Exception(f"Failed to append detailed records: {e}")

    def generate_report(self, filters, output_path=None):
        """
        Generates report with Butlo-dni calculation.
        
        Steps:
        1. Get data from approved files (from database).
        2. Sort: Company -> Product -> Date.
        3. Calculate logic (Next Date, Days, Butlo-dni).
        4. Generate Excel with 3 sheets.
        5. Create Pivot Table using Win32 COM (Isolated Instance).
        
        Args:
            filters: Dictionary with filter criteria
            output_path: Optional custom path for the output file. If None, uses default location.
        """
        from datetime import datetime, date
        import pandas as pd
        from calendar import monthrange
        from core.database_handler import DatabaseHandler
        from config import DATABASE_FILE
        from openpyxl.utils import get_column_letter
        import os
        
        # Get approved files from database instead of file system
        db = DatabaseHandler(DATABASE_FILE)
        approved_records = db.get_all_approved_records()
        print(f"[REPORT] Found {len(approved_records) if approved_records else 0} approved records in database")
        print(f"[REPORT] Filters: {filters}")
        
        if not approved_records:
            raise Exception("Nie znaleziono zatwierdzonych raportów.")
        
        # Build list of file paths from database records
        approved_files = []
        for record in approved_records:
            filepath = record.get('filepath')
            if filepath and os.path.exists(filepath):
                approved_files.append(filepath)
        
        print(f"[REPORT] Found {len(approved_files)} existing approved files")
        if not approved_files:
            raise Exception("Nie znaleziono zatwierdzonych raportów (pliki nie istnieją).")
        
        all_data = []
        
        # --- 1. DATA LOADING (Unchanged) ---
        for file_path in approved_files:
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                
                odbiorca = str(ws['B1'].value).strip() if ws['B1'].value else None
                data_wyst = ws['D1'].value
                nr_dok = self._normalize_invoice_number(str(ws['F1'].value)) if ws['F1'].value else None
                
                if isinstance(data_wyst, str):
                    try:
                        data_wyst = datetime.strptime(data_wyst.strip(), '%d.%m.%Y').date()
                    except: pass
                elif isinstance(data_wyst, datetime):
                    data_wyst = data_wyst.date()

                if filters['mode'] in [2, 3] and filters['company']:
                    if not odbiorca or filters['company'].lower() not in odbiorca.lower():
                        continue

                rows = []
                for r in range(4, ws.max_row + 1):
                    nazwa = ws.cell(row=r, column=2).value
                    stan_po = ws.cell(row=r, column=7).value
                    
                    if not nazwa and stan_po is None: continue
                        
                    row_dict = {
                        'Odbiorca': odbiorca,
                        'NIP': self.extract_nip(odbiorca),
                        'Data wystawienia': data_wyst,
                        'Nr dokumentu': nr_dok,
                        'Nazwa': str(nazwa).strip() if nazwa else "Nieokreślony",
                        'Ilość zamówiona': ws.cell(row=r, column=3).value,
                        'Ilość zwrócona': ws.cell(row=r, column=5).value,
                        'stan poprzedni': ws.cell(row=r, column=6).value,
                        'stan po wymianie': stan_po
                    }
                    rows.append(row_dict)
                
                if rows:
                    all_data.append(pd.DataFrame(rows))
                    
            except Exception as e:
                print(f"Błąd przy pliku {file_path}: {e}")
                continue

        if not all_data:
            raise Exception(f"Brak danych. Przeszukano {len(approved_files)} plików.")

        df = pd.concat(all_data, ignore_index=True)
        print(f"[REPORT] Raw df shape: {df.shape}")
        print(f"[REPORT] Sample data_wystawienia values: {df['Data wystawienia'].head(10).tolist()}")

        # --- 2. CLEANING & FORMATTING (Unchanged) ---
        # Handle both Python date/datetime objects and strings
        def convert_date(x):
            if pd.isna(x):
                return pd.NaT
            # If it's already a datetime object, use it
            if isinstance(x, (datetime, pd.Timestamp)):
                return pd.Timestamp(x)
            # If it's a date object, convert to datetime
            if isinstance(x, date) and not isinstance(x, datetime):
                return pd.Timestamp(x)
            # If it's a string, try parsing with format
            if isinstance(x, str):
                try:
                    return pd.Timestamp(datetime.strptime(x, '%Y-%m-%d'))
                except:
                    return pd.NaT
            return pd.NaT
        
        df['Data wystawienia'] = df['Data wystawienia'].apply(convert_date)
        print(f"[REPORT] After conversion to datetime:")
        print(f"[REPORT] Sample converted values: {df['Data wystawienia'].head(10).tolist()}")
        print(f"[REPORT] Number of NaT values: {df['Data wystawienia'].isna().sum()}")
        
        numeric_cols = ['stan po wymianie', 'stan poprzedni', 'Ilość zamówiona', 'Ilość zwrócona']
        for col in numeric_cols:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '.'), errors='coerce').fillna(0)

        df_all = df.copy()
        df_all['Miesiąc'] = df_all['Data wystawienia'].dt.strftime('%Y-%m')
        print(f"[REPORT] Months in df_all: {df_all['Miesiąc'].unique().tolist()}")
        latest_month = None
        valid_dates = df_all['Data wystawienia'].dropna()
        if not valid_dates.empty:
            latest_month = valid_dates.dt.to_period('M').max().strftime('%Y-%m')
        df_current = None
        if latest_month:
            df_current = df_all[df_all['Miesiąc'] == latest_month].copy()

        df_all = df.copy()
        df_all['Miesiąc'] = df_all['Data wystawienia'].dt.strftime('%Y-%m')
        latest_month = None
        valid_dates = df_all['Data wystawienia'].dropna()
        if not valid_dates.empty:
            latest_month = valid_dates.dt.to_period('M').max().strftime('%Y-%m')
        df_current = None
        if latest_month:
            df_current = df_all[df_all['Miesiąc'] == latest_month].copy()

        print(f"[REPORT] Loaded df_all with {len(df_all)} rows")
        print(f"[REPORT] Available months in data: {df_all['Miesiąc'].unique().tolist() if 'Miesiąc' in df_all.columns else 'No Miesiąc column'}")
        
        # --- 2.5. FILL MISSING MONTHS (Carry-forward) ---
        report_start = None
        report_end = None
        if filters['mode'] == 1 and filters.get('month'):
            year, month = map(int, filters['month'].split('-'))
            report_start = pd.Timestamp(year=year, month=month, day=1)
            report_end = pd.Timestamp(year=year, month=month, day=monthrange(year, month)[1])
        elif filters.get('to_date'):
            report_end = pd.Timestamp(filters['to_date'])
            if filters.get('from_date'):
                report_start = pd.Timestamp(filters['from_date'])
        else:
            report_end = pd.Timestamp(datetime.now().date())

        if report_end is not None:
            min_date = df['Data wystawienia'].min()
            if pd.notna(min_date):
                start_month = pd.Timestamp(year=min_date.year, month=min_date.month, day=1)
                if report_start is not None:
                    start_month = max(start_month, pd.Timestamp(year=report_start.year, month=report_start.month, day=1))
                month_starts = pd.date_range(start=start_month, end=report_end, freq='MS')
                synthetic_rows = []

                for (odbiorca, nazwa), group in df.groupby(['Odbiorca', 'Nazwa']):
                    group_sorted = group.sort_values('Data wystawienia')
                    existing_months = set(group_sorted['Data wystawienia'].dt.strftime('%Y-%m'))

                    for ms in month_starts:
                        month_key = ms.strftime('%Y-%m')
                        if month_key in existing_months:
                            continue
                        prior = group_sorted[group_sorted['Data wystawienia'] < ms]
                        if prior.empty:
                            continue
                        last_row = prior.iloc[-1]

                        synthetic = last_row.copy()
                        synthetic['Data wystawienia'] = ms
                        synthetic['Ilość zamówiona'] = 0
                        synthetic['Ilość zwrócona'] = 0
                        synthetic['stan poprzedni'] = last_row['stan po wymianie']
                        synthetic['stan po wymianie'] = last_row['stan po wymianie']
                        synthetic_rows.append(synthetic)

                if synthetic_rows:
                    df = pd.concat([df, pd.DataFrame(synthetic_rows)], ignore_index=True)

        if filters['mode'] == 1 and filters.get('month'):
            df['Miesiąc_temp'] = df['Data wystawienia'].dt.strftime('%Y-%m')
            print(f"[REPORT] Before month filter: {len(df)} rows, filtering for month={filters['month']}")
            df = df[df['Miesiąc_temp'] == filters['month']].drop(columns=['Miesiąc_temp'])
            print(f"[REPORT] After month filter: {len(df)} rows")
        elif filters['mode'] in [1, 2] and filters.get('from_date') and filters.get('to_date'):
            f_start = pd.Timestamp(filters['from_date'])
            f_end = pd.Timestamp(filters['to_date'])
            print(f"[REPORT] Before date range filter: {len(df)} rows, filtering from {f_start} to {f_end}")
            df = df[(df['Data wystawienia'] >= f_start) & (df['Data wystawienia'] <= f_end)]
            print(f"[REPORT] After date range filter: {len(df)} rows")

        if df.empty:
            available_dates = f"from {df_all['Data wystawienia'].min().date()} to {df_all['Data wystawienia'].max().date()}" if not df_all.empty else "no data"
            error_msg = f"Brak danych po przefiltrowaniu dat. Available data: {available_dates}"
            print(f"[REPORT] ERROR: {error_msg}")
            raise Exception(error_msg)

        # --- 3. BUTLO-DNI LOGIC (Unchanged) ---
        df.sort_values(by=['Odbiorca', 'Nazwa', 'Data wystawienia'], ascending=[True, True, True], inplace=True)
        df['Data następna'] = df.groupby(['Odbiorca', 'Nazwa'])['Data wystawienia'].shift(-1)
        today = pd.Timestamp(datetime.now().date())
        df['Data następna'] = df['Data następna'].fillna(today)
        df['Liczba dni'] = (df['Data następna'] - df['Data wystawienia']).dt.days
        df['Liczba dni'] = df['Liczba dni'].apply(lambda x: x if x >= 0 else 0)
        df['Miesiąc'] = df['Data wystawienia'].dt.strftime('%Y-%m')
        df['Is_first_of_month'] = df.groupby(['Odbiorca', 'Nazwa', 'Miesiąc']).cumcount() == 0
        df['Butlo-dni'] = df.apply(
            lambda row: row['stan poprzedni'] * row['Liczba dni'] if row['Is_first_of_month'] else row['stan po wymianie'] * row['Liczba dni'],
            axis=1
        )

        # --- 4. EXCEL EXPORT (Unchanged) ---
        df = self._fill_missing_nip_from_db(df)
        if df_current is not None:
            df_current = self._fill_missing_nip_from_db(df_current)
        raw_cols = ['Odbiorca', 'NIP', 'Nazwa', 'Nr dokumentu', 'Data wystawienia', 'Ilość zamówiona', 'Ilość zwrócona', 'stan poprzedni', 'stan po wymianie', 'Miesiąc']
        df_raw = df[raw_cols].copy()
        if df_current is not None:
            df_current_raw = df_current[raw_cols].copy()
        else:
            df_current_raw = pd.DataFrame(columns=raw_cols)
        if df_current is not None:
            df_current_raw = df_current[raw_cols].copy()
        else:
            df_current_raw = pd.DataFrame(columns=raw_cols)

        # Ensure dates are stored without time for Excel output
        for frame in (df_raw, df_current_raw):
            if 'Data wystawienia' in frame.columns:
                frame['Data wystawienia'] = pd.to_datetime(frame['Data wystawienia'], errors='coerce').dt.date

        calc_rows = []
        for (odbiorca, nazwa, miesiac), group in df.groupby(['Odbiorca', 'Nazwa', 'Miesiąc']):
            year, month = map(int, miesiac.split('-'))
            first_day = pd.Timestamp(year=year, month=month, day=1)
            last_day = pd.Timestamp(year=year, month=month, day=monthrange(year, month)[1])
            
            group = group.sort_values('Data wystawienia').reset_index(drop=True)
            current_start = first_day
            
            for idx, row in group.iterrows():
                report_date = row['Data wystawienia']
                if idx == 0:
                    days = (report_date - current_start).days
                    stan = row['stan poprzedni']
                    nr_doc = row['Nr dokumentu']
                else:
                    days = (report_date - current_start).days
                    stan = group.loc[idx - 1, 'stan po wymianie']
                    nr_doc = group.loc[idx - 1, 'Nr dokumentu']

                if days > 0:
                    calc_rows.append({
                        'Odbiorca': odbiorca, 'Nazwa': nazwa, 'Nr dokumentu': nr_doc,
                        'Data początkowa': current_start, 'Data końcowa': report_date - pd.Timedelta(days=1),
                        'liczba dni': days, 'Stan': stan,
                        'Butlo-dni': stan * days, 'Miesiąc': miesiac
                    })
                current_start = report_date
            
            last_row = group.iloc[-1]
            days_end = (last_day - current_start).days + 1
            stan_end = last_row['stan po wymianie']
            calc_rows.append({
                'Odbiorca': odbiorca, 'Nazwa': nazwa, 'Nr dokumentu': last_row['Nr dokumentu'],
                'Data początkowa': current_start, 'Data końcowa': last_day, 'liczba dni': days_end, 'Stan': stan_end,
                'Butlo-dni': stan_end * days_end if days_end > 0 else 0, 'Miesiąc': miesiac
            })
        
        df_calc = pd.DataFrame(calc_rows)

        # Remove time from calculation dates for Excel output
        for col in ['Data początkowa', 'Data końcowa']:
            if col in df_calc.columns:
                df_calc[col] = pd.to_datetime(df_calc[col], errors='coerce').dt.date
        
        # Use provided output_path or generate default
        if not output_path:
            output_filename = f"Raport_ButloDni_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            output_path = os.path.join(os.path.dirname(APPROVED_FILE), output_filename)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Summary Logic (needed for pivots, but won't display)
            df_calc_pos = df_calc.copy()
            df_calc_pos['Butlo-dni'] = df_calc_pos['Butlo-dni'].apply(lambda x: max(0, x))
            sum_butlo = df_calc_pos.groupby(['Odbiorca', 'Nazwa', 'Miesiąc'], as_index=False)[['Butlo-dni']].sum()
            sum_rot = df_raw.groupby(['Odbiorca', 'Nazwa', 'Miesiąc'], as_index=False)[['Ilość zwrócona']].sum()
            summary = sum_butlo.merge(sum_rot, on=['Odbiorca', 'Nazwa', 'Miesiąc'], how='outer')
            nip_map = df[['Odbiorca', 'NIP']].drop_duplicates()
            summary = summary.merge(nip_map, on='Odbiorca', how='left')
            summary = summary[['Odbiorca', 'NIP', 'Nazwa', 'Miesiąc', 'Butlo-dni', 'Ilość zwrócona']]
            summary = summary.rename(columns={'Ilość zwrócona': 'rotacja'})
            summary.to_excel(writer, sheet_name='Podsumowanie', index=False)
            
            # Podsumowanie butlodni - same as Obliczenia but with NIP column
            summary_butlodni = df_calc.copy()
            summary_butlodni = summary_butlodni.merge(nip_map, on='Odbiorca', how='left')
            
            # Get Stan bieżący from the last record of each Odbiorca-Nazwa pair (most recent across ALL data)
            df_all_sorted = df_all.sort_values('Data wystawienia')
            stan_biezacy_map = df_all_sorted.groupby(['Odbiorca', 'Nazwa'], as_index=False).tail(1)
            stan_biezacy_map = stan_biezacy_map[['Odbiorca', 'Nazwa', 'stan po wymianie']]
            stan_biezacy_map = stan_biezacy_map.rename(columns={'stan po wymianie': 'Stan bieżący'})
            
            rotacja_map = sum_rot.rename(columns={'Ilość zwrócona': 'rotacja'})
            summary_butlodni = summary_butlodni.merge(rotacja_map, on=['Odbiorca', 'Nazwa', 'Miesiąc'], how='left')
            summary_butlodni = summary_butlodni.merge(stan_biezacy_map, on=['Odbiorca', 'Nazwa'], how='left')
            summary_butlodni = summary_butlodni[['Odbiorca', 'NIP', 'Nazwa', 'Nr dokumentu', 'Data początkowa', 'Data końcowa', 'liczba dni', 'Stan', 'Butlo-dni', 'rotacja', 'Stan bieżący', 'Miesiąc']]
            summary_butlodni.to_excel(writer, sheet_name='Podsumowanie butlodni', index=False)
            
            # Create Rotacja summary - monthly totals with last-day Stan
            rotacja_rows = []
            for (odbiorca, nazwa, miesiac), group in df_calc.groupby(['Odbiorca', 'Nazwa', 'Miesiąc']):
                # Get the last row (end of month) for Stan
                last_row = group.sort_values('Data końcowa').iloc[-1]
                
                # Sum Butlo-dni for the month
                total_butlodni = group['Butlo-dni'].sum()
                
                # Get rotacja for this company/product/month
                rotacja_val = sum_rot[
                    (sum_rot['Odbiorca'] == odbiorca) & 
                    (sum_rot['Nazwa'] == nazwa) & 
                    (sum_rot['Miesiąc'] == miesiac)
                ]['Ilość zwrócona'].sum() if not sum_rot.empty else 0
                
                rotacja_rows.append({
                    'Odbiorca': odbiorca,
                    'Nazwa': nazwa,
                    'Miesiąc': miesiac,
                    'Stan': last_row['Stan'],
                    'Butlo-dni': total_butlodni,
                    'rotacja': rotacja_val
                })
            
            rotacja_summary = pd.DataFrame(rotacja_rows)
            rotacja_summary = rotacja_summary.merge(nip_map, on='Odbiorca', how='left')
            rotacja_summary = rotacja_summary[['Odbiorca', 'NIP', 'Nazwa', 'Miesiąc', 'Stan', 'Butlo-dni', 'rotacja']]
            rotacja_summary.to_excel(writer, sheet_name='Rotacja Source', index=False)
            
            # Create Daily breakdown - one row per day for each Odbiorca-Nazwa pair
            daily_rows = []
            for (odbiorca, nazwa, miesiac), group in df_calc.groupby(['Odbiorca', 'Nazwa', 'Miesiąc']):
                # Parse the month to get the date range
                year, month = map(int, miesiac.split('-'))
                first_day = pd.Timestamp(year=year, month=month, day=1)
                last_day = (first_day + pd.DateOffset(months=1)) - pd.DateOffset(days=1)
                
                # Get NIP for this company
                nip_val = nip_map[nip_map['Odbiorca'] == odbiorca]['NIP'].iloc[0] if len(nip_map[nip_map['Odbiorca'] == odbiorca]) > 0 else None
                
                # Get Stan bieżący from the mapped values (latest state for this company-product pair)
                matching = stan_biezacy_map[(stan_biezacy_map['Odbiorca'] == odbiorca) & (stan_biezacy_map['Nazwa'] == nazwa)]
                stan_biezacy = matching['Stan bieżący'].iloc[0] if len(matching) > 0 else None
                
                # Get actual rotacja transactions for this company/product/month from df_raw
                rotacja_transactions = df_raw[
                    (df_raw['Odbiorca'] == odbiorca) & 
                    (df_raw['Nazwa'] == nazwa) & 
                    (df_raw['Miesiąc'] == miesiac)
                ].copy()
                
                # Convert Data wystawienia to datetime if it's not already
                if not rotacja_transactions.empty:
                    rotacja_transactions['Data wystawienia'] = pd.to_datetime(rotacja_transactions['Data wystawienia'])
                
                # Group rotacja by date (sum all returns that happened on same date)
                # Convert to date only (without time) for matching with current_date
                if not rotacja_transactions.empty:
                    rotacja_by_date = rotacja_transactions.groupby(rotacja_transactions['Data wystawienia'].dt.date)['Ilość zwrócona'].sum().to_dict()
                    total_rotacja = sum(rotacja_by_date.values())
                    print(f"✓ {odbiorca} - {nazwa} - {miesiac}: Found {len(rotacja_by_date)} return dates, total rotacja = {total_rotacja}")
                else:
                    rotacja_by_date = {}
                    print(f"✗ {odbiorca} - {nazwa} - {miesiac}: No return data (df_raw has {len(df_raw)} rows total)")
                
                # Pre-compute Stan values for each day (much faster than nested loop)
                stan_by_date = {}
                for _, row in group.iterrows():
                    date_range_start = pd.Timestamp(row['Data początkowa'])
                    date_range_end = pd.Timestamp(row['Data końcowa'])
                    stan_val = row['Stan']
                    
                    # Fill in stan for each day in the range
                    current = date_range_start
                    while current <= date_range_end:
                        stan_by_date[current.date()] = stan_val
                        current += pd.DateOffset(days=1)
                
                # Create a row for each day in the month
                current_date = first_day
                while current_date <= last_day:
                    date_key = current_date if not hasattr(current_date, 'date') else current_date.date()
                    # Look up stan from pre-computed dictionary
                    stan_on_date = stan_by_date.get(date_key)
                    butlodni_on_date = stan_on_date if stan_on_date is not None else 0
                    
                    # Get rotacja only if there was a return on this specific date
                    rotacja_on_date = rotacja_by_date.get(date_key, 0)
                    
                    daily_rows.append({
                        'Odbiorca': odbiorca,
                        'NIP': nip_val,
                        'Nazwa': nazwa,
                        'Data': date_key,
                        'Stan': stan_on_date,
                        'Butlo-dni': butlodni_on_date,
                        'rotacja': rotacja_on_date,
                        'rotacja miesięczna': total_rotacja,
                        'Stan bieżący': stan_biezacy
                    })
                    
                    current_date += pd.DateOffset(days=1)
            
            daily_df = pd.DataFrame(daily_rows)
            daily_df = daily_df[['Odbiorca', 'NIP', 'Nazwa', 'Data', 'Stan', 'Butlo-dni', 'rotacja', 'rotacja miesięczna', 'Stan bieżący']]
            daily_df.to_excel(writer, sheet_name='Daily Data', index=False)

        # ---------------------------------------------------------
        # PIVOT GENERATION (THREAD-SAFE & ROBUST)
        # ---------------------------------------------------------
        excel = None
        wb_com = None
        
        import pythoncom 
        import win32com.client as win32
        import time
        import gc

        try:
            # 1. Initialize COM for the current thread
            pythoncom.CoInitialize()

            # 2. Give the OS a moment to release the file handle from Pandas
            time.sleep(1.5)
            
            abs_path = os.path.abspath(output_path)
            
            # Constants
            xlDatabase, xlRowField, xlColumnField, xlPageField = 1, 1, 2, 3
            xlSum, xlUp, xlToLeft = -4157, -4162, -4159

            # 3. Create Excel Instance with Error Handling
            try:
                # DispatchEx forces a NEW process, preventing "Zombie" interference
                excel = win32.DispatchEx("Excel.Application")
            except Exception as com_err:
                print(f"Failed to start Excel via DispatchEx: {com_err}")
                # Fallback
                excel = win32.GetActiveObject("Excel.Application")
            
            excel.Visible = False
            excel.DisplayAlerts = False 

            # 4. Open Workbook with a retry loop
            for attempt in range(3):
                try:
                    wb_com = excel.Workbooks.Open(abs_path)
                    break
                except Exception as e:
                    if attempt == 2: raise e
                    time.sleep(1)

            ws_src = wb_com.Sheets('Podsumowanie')
            
            # 5. Define Range for first source
            lastRow = ws_src.Cells(ws_src.Rows.Count, 1).End(xlUp).Row
            lastCol = ws_src.Cells(1, ws_src.Columns.Count).End(xlToLeft).Column
            src_range = ws_src.Range(ws_src.Cells(1, 1), ws_src.Cells(lastRow, lastCol))

            # Define Range for second source
            ws_src2 = wb_com.Sheets('Podsumowanie butlodni')
            lastRow2 = ws_src2.Cells(ws_src2.Rows.Count, 1).End(xlUp).Row
            lastCol2 = ws_src2.Cells(1, ws_src2.Columns.Count).End(xlToLeft).Column
            src_range2 = ws_src2.Range(ws_src2.Cells(1, 1), ws_src2.Cells(lastRow2, lastCol2))
            
            # 6. Create rotacja_1 pivot from Podsumowanie butlodni
            # Delete if exists
            try:
                wb_com.Sheets("rotacja_1").Delete()
            except:
                pass
            
            ws_pivot1 = wb_com.Sheets.Add()
            ws_pivot1.Name = "rotacja_1"
            
            pc1 = wb_com.PivotCaches().Create(SourceType=xlDatabase, SourceData=src_range2)
            pt1 = pc1.CreatePivotTable(TableDestination=ws_pivot1.Range("A3"), TableName="Rotacja1Pivot")
            
            pt1.PivotFields('Miesiąc').Orientation = xlPageField
            pt1.PivotFields('NIP').Orientation = xlPageField
            pt1.PivotFields('Odbiorca').Orientation = xlPageField
            
            row_fields_3 = ['Nazwa', 'Stan bieżący', 'Stan']
            for i, f in enumerate(row_fields_3, start=1):
                pf = pt1.PivotFields(f)
                pf.Orientation = xlRowField
                pf.Position = i
                pf.Subtotals = [False] * 12
            
            f1 = pt1.AddDataField(pt1.PivotFields('Butlo-dni'), "Suma Butlo-dni", xlSum)
            f1.NumberFormat = "0.00"
            f2 = pt1.AddDataField(pt1.PivotFields('rotacja'), "Sum of rotacja", xlSum)
            f2.NumberFormat = "0"
            
            try:
                pt1.DataPivotField.Orientation = xlColumnField
            except:
                pass
            
            pt1.RowAxisLayout(1)
            ws_pivot1.Activate()
            wb_com.Save()
            print("✓ rotacja_1 pivot table created successfully.")
            
            # 7. Create Raport butlodni pivot from Podsumowanie butlodni
            try:
                wb_com.Sheets("Raport butlodni").Delete()
            except:
                pass
            
            ws_pivot2 = wb_com.Sheets.Add()
            ws_pivot2.Name = 'Raport butlodni'
            
            pc2 = wb_com.PivotCaches().Create(SourceType=xlDatabase, SourceData=src_range2)
            pt2 = pc2.CreatePivotTable(TableDestination=ws_pivot2.Range("A3"), TableName="RaportButlodni")
            
            pt2.ManualUpdate = True
            pt2.PivotFields('Odbiorca').Orientation = xlPageField
            pt2.PivotFields('NIP').Orientation = xlPageField
            pt2.PivotFields('Miesiąc').Orientation = xlPageField
            
            row_fields = ['Nazwa', 'Data początkowa', 'Data końcowa', 'Stan']
            for i, f in enumerate(row_fields, start=1):
                pf = pt2.PivotFields(f)
                pf.Orientation = xlRowField
                pf.Position = i
                pf.Subtotals = [False] * 12
                pf.LayoutForm = 1
                try:
                    pf.Width = 50
                except:
                    pass
            
            pt2.RowAxisLayout(1)
            pt2.RepeatAllLabels(2)
            pt2.ManualUpdate = False
            
            f3 = pt2.AddDataField(pt2.PivotFields('Butlo-dni'), "Suma Butlo-dni", xlSum)
            f3.NumberFormat = "0.00"
            f4 = pt2.AddDataField(pt2.PivotFields('rotacja'), "Suma rotacja", xlSum)
            f4.NumberFormat = "0"
            
            try:
                pt2.DataPivotField.Orientation = xlColumnField
            except:
                pass
            
            pt2.RowAxisLayout(1)
            pt2.RefreshTable()
            
            # Set column widths to auto-fit instead of fixed 80
            for col_idx in range(1, 20):
                try:
                    ws_pivot2.Columns(col_idx).ColumnWidth = 18
                except:
                    pass
            
            ws_pivot2.Activate()
            wb_com.Save()
            print("✓ Raport butlodni pivot table created successfully.")
            
            # 8. Create Rotacja pivot from Rotacja Source
            try:
                wb_com.Sheets("Rotacja").Delete()
            except:
                pass
            
            ws_src3 = wb_com.Sheets('Rotacja Source')
            lastRow3 = ws_src3.Cells(ws_src3.Rows.Count, 1).End(xlUp).Row
            lastCol3 = ws_src3.Cells(1, ws_src3.Columns.Count).End(xlToLeft).Column
            src_range3 = ws_src3.Range(ws_src3.Cells(1, 1), ws_src3.Cells(lastRow3, lastCol3))
            
            ws_pivot3 = wb_com.Sheets.Add()
            ws_pivot3.Name = 'Rotacja'
            
            pc3 = wb_com.PivotCaches().Create(SourceType=xlDatabase, SourceData=src_range3)
            pt3 = pc3.CreatePivotTable(TableDestination=ws_pivot3.Range("A3"), TableName="RotacjaPivot")
            
            pt3.PivotFields('Miesiąc').Orientation = xlPageField
            pt3.PivotFields('NIP').Orientation = xlPageField
            pt3.PivotFields('Odbiorca').Orientation = xlPageField
            
            # Row fields: just Nazwa and Stan
            pt3.PivotFields('Nazwa').Orientation = xlRowField
            pt3.PivotFields('Nazwa').Position = 1
            pt3.PivotFields('Nazwa').Subtotals = [False] * 12
            
            pt3.PivotFields('Stan').Orientation = xlRowField
            pt3.PivotFields('Stan').Position = 2
            pt3.PivotFields('Stan').Subtotals = [False] * 12
            
            # Add data fields
            f6 = pt3.AddDataField(pt3.PivotFields('Butlo-dni'), "Suma Butlo-dni", xlSum)
            f6.NumberFormat = "0.00"
            f7 = pt3.AddDataField(pt3.PivotFields('rotacja'), "Sum of rotacja", xlSum)
            f7.NumberFormat = "0"
            
            pt3.DataPivotField.Orientation = xlColumnField
            pt3.RowAxisLayout(1)
            ws_pivot3.Activate()
            wb_com.Save()
            print("✓ Rotacja pivot table created successfully.")
            
            # 8.5. Create Rotacja_2 pivot from Daily Data
            try:
                wb_com.Sheets("Rotacja_2").Delete()
            except:
                pass
            
            ws_src_daily = wb_com.Sheets('Daily Data')
            lastRow_daily = ws_src_daily.Cells(ws_src_daily.Rows.Count, 1).End(xlUp).Row
            lastCol_daily = ws_src_daily.Cells(1, ws_src_daily.Columns.Count).End(xlToLeft).Column
            src_range_daily = ws_src_daily.Range(ws_src_daily.Cells(1, 1), ws_src_daily.Cells(lastRow_daily, lastCol_daily))
            
            ws_pivot_daily = wb_com.Sheets.Add()
            ws_pivot_daily.Name = 'Rotacja_2'
            
            pc_daily = wb_com.PivotCaches().Create(SourceType=xlDatabase, SourceData=src_range_daily)
            pt_daily = pc_daily.CreatePivotTable(TableDestination=ws_pivot_daily.Range("A3"), TableName="Rotacja2Pivot")
            
            # Page fields (filters)
            pt_daily.PivotFields('Odbiorca').Orientation = xlPageField
            pt_daily.PivotFields('NIP').Orientation = xlPageField
            pt_daily.PivotFields('Data').Orientation = xlPageField
            
            # Row fields: Nazwa, Stan, Stan bieżący
            pt_daily.PivotFields('Nazwa').Orientation = xlRowField
            pt_daily.PivotFields('Nazwa').Position = 1
            pt_daily.PivotFields('Nazwa').Subtotals = [False] * 12
            
            pt_daily.PivotFields('Stan').Orientation = xlRowField
            pt_daily.PivotFields('Stan').Position = 2
            pt_daily.PivotFields('Stan').Subtotals = [False] * 12
            
            pt_daily.PivotFields('Stan bieżący').Orientation = xlRowField
            pt_daily.PivotFields('Stan bieżący').Position = 3
            pt_daily.PivotFields('Stan bieżący').Subtotals = [False] * 12
            
            # Add data fields
            f_rot = pt_daily.AddDataField(pt_daily.PivotFields('rotacja'), "Sum of rotacja", xlSum)
            f_rot.NumberFormat = "0"
            f_rot_miesiac = pt_daily.AddDataField(pt_daily.PivotFields('rotacja miesięczna'), "Sum of rotacja miesięczna Do usu", xlSum)
            f_rot_miesiac.NumberFormat = "0"
            
            try:
                pt_daily.DataPivotField.Orientation = xlColumnField
            except Exception as e:
                print(f"Warning: Could not set DataPivotField orientation: {e}")
            
            pt_daily.RowAxisLayout(1)
            pt_daily.RefreshTable()  # Refresh to ensure all fields display
            ws_pivot_daily.Activate()
            wb_com.Save()
            print("✓ Rotacja_2 pivot table created successfully.")
            
            # 9. Hide source sheets
            wb_com.Sheets('Podsumowanie').Visible = False
            wb_com.Sheets('Podsumowanie butlodni').Visible = False
            wb_com.Sheets('Rotacja Source').Visible = False
            wb_com.Sheets('Daily Data').Visible = False
            wb_com.Save()
            print("✓ Source sheets hidden.")

        except Exception as e:
            print(f"Critical Pivot Error: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            # 8. Aggressive Cleanup to prevent memory leaks/zombies
            if wb_com:
                try: wb_com.Close(SaveChanges=True)
                except: pass
                del wb_com
            
            if excel:
                try: excel.Quit()
                except: pass
                del excel
            
            # Force Windows to clear COM hooks
            gc.collect()
            pythoncom.CoUninitialize()

        return output_path

    def update_approved_date(self, filename, new_date):
        """
        Update the date for a specific filename in database.
        Used when the date in the Excel file is edited after approval.
        """
        try:
            success = self.db.update_approved_date(filename, new_date)
            if success:
                print(f"✓ Updated date in database for '{filename}' to '{new_date}'")
            else:
                print(f"Warning: Record not found for '{filename}'")
        except Exception as e:
            print(f"Warning: Could not update date in database: {e}")

        import pandas as pd
        from calendar import monthrange
        from core.file_manager import FileManager
        from openpyxl.utils import get_column_letter
        
        file_manager = FileManager()
        approved_files = file_manager.get_approved_reports()
        
        if not approved_files:
            raise Exception("Nie znaleziono zatwierdzonych raportów.")
        
        all_data = []
        
        # --- 1. DATA LOADING (Unchanged) ---
        for file_path in approved_files:
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                
                odbiorca = str(ws['B1'].value).strip() if ws['B1'].value else None
                data_wyst = ws['D1'].value
                nr_dok = self._normalize_invoice_number(str(ws['F1'].value)) if ws['F1'].value else None
                
                if isinstance(data_wyst, str):
                    try:
                        data_wyst = datetime.strptime(data_wyst.strip(), '%d.%m.%Y').date()
                    except:
                        try:
                            data_wyst = datetime.strptime(data_wyst.strip(), '%d.%m.%y').date()
                        except: pass
                elif isinstance(data_wyst, datetime):
                    data_wyst = data_wyst.date()

                if filters['mode'] in [2, 3] and filters['company']:
                    if not odbiorca or filters['company'].lower() not in odbiorca.lower():
                        continue

                rows = []
                for r in range(4, ws.max_row + 1):
                    nazwa = ws.cell(row=r, column=2).value
                    stan_po = ws.cell(row=r, column=7).value
                    
                    if not nazwa and stan_po is None: continue
                        
                    row_dict = {
                        'Odbiorca': odbiorca,
                        'NIP': self.extract_nip(odbiorca),
                        'Data wystawienia': data_wyst,
                        'Nr dokumentu': nr_dok,
                        'Nazwa': str(nazwa).strip() if nazwa else "Nieokreślony",
                        'Ilość zamówiona': ws.cell(row=r, column=3).value,
                        'Ilość zwrócona': ws.cell(row=r, column=5).value,
                        'stan poprzedni': ws.cell(row=r, column=6).value,
                        'stan po wymianie': stan_po
                    }
                    rows.append(row_dict)
                
                if rows:
                    all_data.append(pd.DataFrame(rows))
                    
            except Exception as e:
                print(f"Błąd przy pliku {file_path}: {e}")
                continue

        if not all_data:
            raise Exception(f"Brak danych. Przeszukano {len(approved_files)} plików.")

        df = pd.concat(all_data, ignore_index=True)

        # --- 2. CLEANING & FORMATTING (Unchanged) ---
        df['Data wystawienia'] = pd.to_datetime(df['Data wystawienia'], errors='coerce')
        numeric_cols = ['stan po wymianie', 'stan poprzedni', 'Ilość zamówiona', 'Ilość zwrócona']
        for col in numeric_cols:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '.'), errors='coerce').fillna(0)

        if filters['mode'] == 1 and filters.get('month'):
            df['Miesiąc_temp'] = df['Data wystawienia'].dt.strftime('%Y-%m')
            df = df[df['Miesiąc_temp'] == filters['month']].drop(columns=['Miesiąc_temp'])
        elif filters['mode'] in [1, 2] and filters.get('from_date') and filters.get('to_date'):
            f_start = pd.Timestamp(filters['from_date'])
            f_end = pd.Timestamp(filters['to_date'])
            df = df[(df['Data wystawienia'] >= f_start) & (df['Data wystawienia'] <= f_end)]

        if df.empty:
            raise Exception("Brak danych po przefiltrowaniu dat.")

        # --- 3. BUTLO-DNI LOGIC (Unchanged) ---
        df.sort_values(by=['Odbiorca', 'Nazwa', 'Data wystawienia'], ascending=[True, True, True], inplace=True)
        df['Data następna'] = df.groupby(['Odbiorca', 'Nazwa'])['Data wystawienia'].shift(-1)
        today = pd.Timestamp(datetime.now().date())
        df['Data następna'] = df['Data następna'].fillna(today)
        df['Liczba dni'] = (df['Data następna'] - df['Data wystawienia']).dt.days
        df['Liczba dni'] = df['Liczba dni'].apply(lambda x: x if x >= 0 else 0)
        df['Miesiąc'] = df['Data wystawienia'].dt.strftime('%Y-%m')
        df['Is_first_of_month'] = df.groupby(['Odbiorca', 'Nazwa', 'Miesiąc']).cumcount() == 0
        df['Butlo-dni'] = df.apply(
            lambda row: row['stan poprzedni'] * row['Liczba dni'] if row['Is_first_of_month'] else row['stan po wymianie'] * row['Liczba dni'],
            axis=1
        )

        # --- 4. EXCEL EXPORT (Unchanged) ---
        df = self._fill_missing_nip_from_db(df)
        raw_cols = ['Odbiorca', 'NIP', 'Nazwa', 'Nr dokumentu', 'Data wystawienia', 'Ilość zamówiona', 'Ilość zwrócona', 'stan poprzedni', 'stan po wymianie', 'Miesiąc']
        df_raw = df[raw_cols].copy()

        calc_rows = []
        for (odbiorca, nazwa, miesiac), group in df.groupby(['Odbiorca', 'Nazwa', 'Miesiąc']):
            year, month = map(int, miesiac.split('-'))
            first_day = pd.Timestamp(year=year, month=month, day=1)
            last_day = pd.Timestamp(year=year, month=month, day=monthrange(year, month)[1])
            
            group = group.sort_values('Data wystawienia').reset_index(drop=True)
            current_start = first_day
            
            for idx, row in group.iterrows():
                report_date = row['Data wystawienia']
                if idx == 0:
                    days = (report_date - current_start).days + 1
                    stan = row['stan poprzedni']
                else:
                    days = (report_date - current_start).days
                    stan = group.loc[idx - 1, 'stan po wymianie']
                
                calc_rows.append({
                    'Odbiorca': odbiorca, 'Nazwa': nazwa, 'Nr dokumentu': row['Nr dokumentu'] if idx == 0 else group.loc[idx - 1, 'Nr dokumentu'],
                    'Data początkowa': current_start, 'Data końcowa': report_date, 'liczba dni': days, 'Stan': stan,
                    'Butlo-dni': stan * days if days > 0 else 0, 'Miesiąc': miesiac
                })
                current_start = report_date
            
            last_row = group.iloc[-1]
            days_end = (last_day - current_start).days
            stan_end = last_row['stan po wymianie']
            calc_rows.append({
                'Odbiorca': odbiorca, 'Nazwa': nazwa, 'Nr dokumentu': last_row['Nr dokumentu'],
                'Data początkowa': current_start, 'Data końcowa': last_day, 'liczba dni': days_end, 'Stan': stan_end,
                'Butlo-dni': stan_end * days_end if days_end > 0 else 0, 'Miesiąc': miesiac
            })
        
        df_calc = pd.DataFrame(calc_rows)
        
        # Use provided output_path or generate default
        if not output_path:
            output_filename = f"Raport_ButloDni_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            output_path = os.path.join(os.path.dirname(APPROVED_FILE), output_filename)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_raw.to_excel(writer, sheet_name='Dane Szczegółowe', index=False)
            df_current_raw.to_excel(writer, sheet_name='bierzący miesiąc', index=False)
            df_calc.to_excel(writer, sheet_name='Obliczenia', index=False)
            
            # Summary Logic
            df_calc_pos = df_calc.copy()
            df_calc_pos['Butlo-dni'] = df_calc_pos['Butlo-dni'].apply(lambda x: max(0, x))
            sum_butlo = df_calc_pos.groupby(['Odbiorca', 'Nazwa', 'Miesiąc'], as_index=False)[['Butlo-dni']].sum()
            sum_rot = df_raw.groupby(['Odbiorca', 'Nazwa', 'Miesiąc'], as_index=False)[['Ilość zwrócona']].sum()
            summary = sum_butlo.merge(sum_rot, on=['Odbiorca', 'Nazwa', 'Miesiąc'], how='outer')
            nip_map = df[['Odbiorca', 'NIP']].drop_duplicates()
            summary = summary.merge(nip_map, on='Odbiorca', how='left')
            summary = summary[['Odbiorca', 'NIP', 'Nazwa', 'Miesiąc', 'Butlo-dni', 'Ilość zwrócona']]
            summary = summary.rename(columns={'Ilość zwrócona': 'rotacja'})
            summary.to_excel(writer, sheet_name='Podsumowanie', index=False)
            
            # Date formatting and auto-width
            for sn in ['Dane Szczegółowe', 'Obliczenia', 'bierzący miesiąc']:
                ws = writer.sheets[sn]
                for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
                     # Simple logic to format date columns (col 4 in Raw, col 4,5 in Calc)
                    if sn in ['Dane Szczegółowe', 'bierzący miesiąc']:
                        r[3].number_format = 'dd.mm.yyyy'
                    else: 
                        r[3].number_format = 'dd.mm.yyyy'
                        r[4].number_format = 'dd.mm.yyyy'
            
            for sn in writer.sheets:
                ws = writer.sheets[sn]
                for col in ws.columns:
                    width = 10
                    for cell in col:
                        try: width = max(width, len(str(cell.value)))
                        except: pass
                    ws.column_dimensions[get_column_letter(col[0].column)].width = width + 2

        # ---------------------------------------------------------
        # ROBUST PIVOT GENERATION (FIXED SECOND-RUN ISSUE)
        # ---------------------------------------------------------
        excel = None
        wb_com = None
        try:
            import win32com.client as win32
            
            abs_path = os.path.abspath(output_path)
            
            # Constants
            xlDatabase = 1
            xlRowField = 1
            xlColumnField = 2
            xlPageField = 3
            xlSum = -4157
            xlUp = -4162
            xlToLeft = -4159

            # FORCE NEW INSTANCE: Use DispatchEx to ensure isolation from previous runs or zombies
            # This prevents attaching to a "dead" excel process from a previous run
            try:
                excel = win32.DispatchEx("Excel.Application")
            except:
                # Fallback if DispatchEx fails
                excel = win32.Dispatch("Excel.Application")

            excel.Visible = False
            excel.DisplayAlerts = False 

            wb_com = excel.Workbooks.Open(abs_path)
            ws_src = wb_com.Sheets('Podsumowanie')
            
            lastRow = ws_src.Cells(ws_src.Rows.Count, 1).End(xlUp).Row
            lastCol = ws_src.Cells(1, ws_src.Columns.Count).End(xlToLeft).Column
            src_range = ws_src.Range(ws_src.Cells(1, 1), ws_src.Cells(lastRow, lastCol))

            ws_pivot = wb_com.Sheets.Add()
            ws_pivot.Name = 'Pivot'

            pc = wb_com.PivotCaches().Create(SourceType=xlDatabase, SourceData=src_range)
            pt = pc.CreatePivotTable(TableDestination=ws_pivot.Range("A3"), TableName="RaportPivot")

            # Filters
            pt.PivotFields('Odbiorca').Orientation = xlPageField
            pt.PivotFields('NIP').Orientation = xlPageField
            pt.PivotFields('Miesiąc').Orientation = xlPageField 

            # Rows
            pt.PivotFields('Nazwa').Orientation = xlRowField

            # Values (Sums)
            df1 = pt.AddDataField(pt.PivotFields('Butlo-dni'), "Suma Butlo-dni", xlSum)
            df1.NumberFormat = "0.00"
            df2 = pt.AddDataField(pt.PivotFields('rotacja'), "Suma rotacja", xlSum)
            df2.NumberFormat = "0"

            # rotacja_1 pivot table from Podsumowanie butlodni
            ws_src2 = wb_com.Sheets('Podsumowanie butlodni')
            lastRow2 = ws_src2.Cells(ws_src2.Rows.Count, 1).End(xlUp).Row
            lastCol2 = ws_src2.Cells(1, ws_src2.Columns.Count).End(xlToLeft).Column
            src_range2 = ws_src2.Range(ws_src2.Cells(1, 1), ws_src2.Cells(lastRow2, lastCol2))

            try:
                wb_com.Sheets('rotacja_1').Delete()
            except Exception:
                pass
            ws_pivot2 = wb_com.Sheets.Add()
            ws_pivot2.Name = 'rotacja_1'

            pc2 = wb_com.PivotCaches().Create(SourceType=xlDatabase, SourceData=src_range2)
            pt2 = pc2.CreatePivotTable(TableDestination=ws_pivot2.Range("A3"), TableName="Rotacja1Pivot")

            pt2.PivotFields('Miesiąc').Orientation = xlPageField
            pt2.PivotFields('NIP').Orientation = xlPageField
            pt2.PivotFields('Odbiorca').Orientation = xlPageField

            row_fields_2 = ['Nazwa', 'Stan bieżący', 'Stan']
            for i, f in enumerate(row_fields_2, start=1):
                pf = pt2.PivotFields(f)
                pf.Orientation = xlRowField
                pf.Position = i
                pf.Subtotals = [False] * 12

            g1 = pt2.AddDataField(pt2.PivotFields('Butlo-dni'), "Suma Butlo-dni", xlSum)
            g1.NumberFormat = "0.00"
            g2 = pt2.AddDataField(pt2.PivotFields('rotacja'), "Sum of rotacja", xlSum)
            g2.NumberFormat = "0"

            try:
                pt2.DataPivotField.Orientation = xlColumnField
            except Exception:
                pass
            pt2.RowAxisLayout(1)

            ws_pivot2.Activate()

            # Columns (Values)
            pt.DataPivotField.Orientation = xlColumnField
            pt.RowAxisLayout(1)
            
            ws_pivot.Activate()
            
            wb_com.Save()
            print("Pivot table created successfully.")

        except Exception as e:
            print(f"Pivot creation failed: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            # ROBUST CLEANUP
            # Ensure we close and quit regardless of success/fail
            if wb_com:
                try:
                    wb_com.Close(SaveChanges=True)
                except: pass
                wb_com = None
            
            if excel:
                try:
                    excel.Quit()
                except: pass
                # Explicitly delete the object to release COM handle
                del excel 

        return output_path

    def delete_approved_record(self, filename):
        """
        Delete a record from database by filename.
        Also deletes corresponding rows from reporting data.
        """
        try:
            # Delete from database
            self.db.delete_reporting_data_by_filename(filename)
            self.db.delete_approved_record(filename)
        except Exception as e:
            print(f"Error deleting from database: {e}")

    def update_approved_date(self, filename, new_date):
        """
        Update the date for a specific filename in database.
        Used when the date in the Excel file is edited after approval.
        """
        try:
            success = self.db.update_approved_date(filename, new_date)
            if success:
                print(f"✓ Updated date in database for '{filename}' to '{new_date}'")
            else:
                print(f"Warning: Record not found for '{filename}'")
        except Exception as e:
            print(f"Warning: Could not update date in database: {e}")
