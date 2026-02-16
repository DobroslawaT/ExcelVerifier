from openpyxl import load_workbook
import glob
import os

files = glob.glob(r'Reports/Zatwierdzone/**/*.xlsx', recursive=True)
print(f"Found {len(files)} files in Zatwierdzone folder")
for f in files:
    try:
        wb = load_workbook(f, data_only=True)
        ws = wb.active
        d1 = ws['D1'].value
        # Try to parse the date
        from datetime import datetime
        if isinstance(d1, str):
            try:
                parsed = datetime.strptime(d1.strip(), '%d.%m.%y')
                print(f"\nFile: {os.path.basename(f)}")
                print(f"D1: {d1} -> {parsed.strftime('%Y-%m-%d')}")
            except:
                print(f"\nFile: {os.path.basename(f)}")
                print(f"D1: {d1} (could not parse)")
        else:
            print(f"\nFile: {os.path.basename(f)}")
            print(f"D1: {d1} (type: {type(d1).__name__})")
        wb.close()
    except Exception as e:
        print(f"Error reading {os.path.basename(f)}: {e}")
