from openpyxl import load_workbook
import glob
import os

files = glob.glob(r'Reports/Zatwierdzone/**/*.xlsx', recursive=True)
print(f"Found {len(files)} files")
for f in files[:1]:
    try:
        wb = load_workbook(f)
        ws = wb.active
        print(f"File: {os.path.basename(f)}")
        print(f"D1 (Data wystawienia): {ws['D1'].value}")
        print(f"Type: {type(ws['D1'].value)}")
        wb.close()
    except Exception as e:
        print(f"Error: {e}")
